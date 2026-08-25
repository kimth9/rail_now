#!/usr/bin/env python3
"""서울지하철 6호선(서울교통공사) 시각표 xlsx -> 정규화 테이블
(parse_seoul5.py와 같은 구조 — "역=행(명시적 도착/출발 라벨), 열차=열")

평일/휴일이 완전히 별개 파일 — daytype을 파일명이 아니라 인자로 받는다.
평일 원본은 역명 글자 사이에 공백이 끼어있다(`화 랑 대`, `신 내`) — clean()에서
공백을 전부 제거해 정규화. 평일 원본에만 "연결 열차"(다음 열차번호) 헤더 행 있음.
"""
import csv
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/240801_6호선 평일 열차운행 시각표(240801).xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_seoul6"
DAYTYPE = sys.argv[3] if len(sys.argv) > 3 else "평일"

TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")
NOT_A_STATION = {"연결열차", "연결열번"}
# 휴일 원본(2023-10-07)에만 남은 절삭·분리 표기 — 평일 원본(2024-08-01, 더 최신)과
# 대조해 확정. 응암은 종착역이라 휴일 원본에 상행/하행 두 표기로 갈라져 있었는데
# 평일 원본은 하나로 합쳐놨다 — 더 최신 쪽을 기준으로 정규화.
RENAME = {
    "디지털미디어": "디지털미디어시티",
    "효창공원": "효창공원앞",
    "응암(상)": "응암",
    "응암(하)": "응암",
}


def clean(v):
    return str(v).strip().replace(" ", "") if v is not None else ""


def to_sec(v):
    if isinstance(v, str):
        m = TIME_RE.match(v.strip())
        if m:
            h, mi, s = (int(x) for x in m.groups())
            return h * 3600 + mi * 60 + s
    return None


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times = [], []
    warnings = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        direction = "up" if "상행" in sheet else "down"

        header_row = next(r for r in range(1, 8) if clean(ws.cell(r, 1).value) == "열차번호")
        meta_labels = {"출발역", "도착역"} | NOT_A_STATION
        link_row = None
        r = header_row + 1
        while clean(ws.cell(r, 1).value) in meta_labels:
            if clean(ws.cell(r, 1).value) in NOT_A_STATION:
                link_row = r
            r += 1
        first_station_row = r

        row_info = []
        current = None
        for rr in range(first_station_row, ws.max_row + 1):
            label = clean(ws.cell(rr, 1).value)
            if label:
                if label in NOT_A_STATION:
                    link_row = link_row or rr
                    current = None
                    continue
                current = RENAME.get(label, label)
                if current not in stops:
                    stops[current] = f"S6M{len(stops) + 1:04d}"
            kind = clean(ws.cell(rr, 2).value)
            if kind in ("도착", "출발") and current is not None:
                row_info.append((rr, current, kind))

        for c in range(3, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            i = 0
            while i < len(row_info):
                name = row_info[i][1]
                entry = {"name": name, "arr": None, "dep": None}
                while i < len(row_info) and row_info[i][1] == name:
                    rr, _, kind = row_info[i]
                    sec = to_sec(ws.cell(rr, c).value)
                    if kind == "도착":
                        entry["arr"] = sec
                    else:
                        entry["dep"] = sec
                    i += 1
                if entry["arr"] is not None or entry["dep"] is not None:
                    raw.append(entry)

            if len(raw) < 2:
                if raw:
                    warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for e in raw:
                for k in ("arr", "dep"):
                    if e[k] is None:
                        continue
                    adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {e['name']} 시각 역행")
                    e[k] = adj
                    prev = adj

            trip_id = f"{DAYTYPE}#{sheet}#{train_no}"
            for seq, e in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif e["arr"] is not None and e["dep"] is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[e["name"]],
                    "arr_sec": e["arr"] if e["arr"] is not None else "",
                    "dep_sec": e["dep"] if e["dep"] is not None else "",
                    "stop_type": kind,
                })
            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "SEOUL6", "line_name": "6호선",
                "formation": "전동차", "direction": direction,
                "service_id": DAYTYPE,
                "origin": stops[raw[0]["name"]],
                "destination": stops[raw[-1]["name"]],
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko"])
        for name, sid in stops.items():
            w.writerow([sid, name])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
