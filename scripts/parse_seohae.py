#!/usr/bin/env python3
"""서해선(전동, 대곡~일산 직결, 원시행) 시각표 xlsx -> 정규화 테이블
(parse_gyeongui.py와 동일 스키마 — 3행 헤더(시발역/종착역/열차번호) + 역당
2행(도착/출발) + '연계열번' 꼬리 행)

입력 구조 (관측):
  - 시트 = 서해선 상행(평일) / 하행(평일) / 상행(휴일) / 하행(휴일)
  - 역 목록이 원시~일산(21개 라벨)까지만 실제 시각이 있고, 그 뒤로 탄현~판문
    (경의선 문산 방면 12개 역)이 전부 빈 셀로 남아있다 — 이 파일이 경의선 파일과
    같은 템플릿을 재사용해 만들어지면서 꼬리 행을 안 지운 것으로 보인다. 시각이
    하나도 없는 역은 애초에 stops에 등록하지 않는다(경의선 파서처럼 "미사용 역"으로
    남겨두는 대신 원천적으로 제외 — 어차피 경의선 파서가 이미 문산 계열 역을 등록함).
  - 노선명 충돌: '일반열차'(무궁화·ITX-마음)에도 서화성~홍성 구간을 도는 별개의
    '서해선'이 이미 등록되어 있다(동해선과 같은 패턴). build_db.py에서
    line_name에 '(전동)' 접미사를 붙여 구분한다.

역명 확정 근거 — 나무위키 "서해선/역 목록"의 대곡~원시 구간표(대곡→원시 순서,
17개 역, 대곡~일산 구간은 경의선과 중복이므로 본선표에 없고 상단 서술로 확인)와
행 순서를 1:1 대조해 전부 확정. 대부분은 기존에 확인된 "신" 접두사 패턴
(신인천→인천, 신수원→수원 등 — 코레일이 다른 노선의 동명역과 구분하려고 붙이는
내부 표기)과 3글자 절삭이 섞여 있다:
  - 신초지 -> 초지 (물리적으로 안산선·수인분당선 초지역과 동일역, "신" 접두사로 구분)
  - 신신현 -> 신현, 신신천 -> 신천 ("신"+2자 실명 패턴 그대로)
  - 신소사 -> 소사 (물리적으로 1호선/경인선 소사역과 동일역)
  - 시흥능/시흥청/시흥대/부천종 -> 시흥능곡/시흥시청/시흥대야/부천종합운동장 (단순 3글자 절삭)
  - 신김포 -> 김포공항: 나무위키 역목록 표에서 원종↔능곡 사이에 오는 역은 김포공항이
    유일 — 위치 대조로 확정. 다만 이 하나만 "신"+2자 패턴이 아니라 "신"+절삭(김포)이
    겹친 이례적 축약형이라 별도로 명시해 둔다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/서해선_전동열차_20260420부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_seohae"

SHEETS = {
    "서해선 상행(평일)": ("평일", "up"), "서해선 하행(평일)": ("평일", "down"),
    "서해선 상행(휴일)": ("휴일", "up"), "서해선 하행(휴일)": ("휴일", "down"),
}

NOT_A_STATION = {"연계열번"}

STATION_ALIAS = {
    "신초지": ("초지", "✔"),
    "시흥능": ("시흥능곡", "✔"),
    "시흥청": ("시흥시청", "✔"),
    "신신현": ("신현", "✔"),
    "신신천": ("신천", "✔"),
    "시흥대": ("시흥대야", "✔"),
    "신소사": ("소사", "✔"),
    "부천종": ("부천종합운동장", "✔"),
    "신김포": ("김포공항", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label or label in NOT_A_STATION:
                continue
            # 탄현~판문(경의선 문산 방면) 꼬리 행 — 전 열차 시각이 비어있으면 애초에
            # 이 노선의 실제 정차역이 아니라고 보고 등록하지 않는다.
            has_data = any(is_time(ws.cell(r, c).value) or is_time(ws.cell(r + 1, c).value)
                           for c in range(2, ws.max_column + 1))
            if not has_data:
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S6{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        link_row = next((r for r in range(header_row + 1, ws.max_row + 1)
                         if "연계" in clean(ws.cell(r, 1).value)), None)

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "SEOHAE", "line_name": "서해선(전동)",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops[raw[0][0]],
                "destination": stops[raw[-1][0]],
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
