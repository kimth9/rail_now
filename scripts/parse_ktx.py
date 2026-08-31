#!/usr/bin/env python3
"""KTX 시각표 xlsx -> GTFS 유사 정규화 테이블 (stops/trips/stop_times/calendar)

입력 구조 (관측):
  - 시트 1개 = 노선 1개 (9개 노선)
  - 시트당 블록 2개: 하행(왼쪽) / 상행(오른쪽), 빈 열로 구분
  - 헤더 3행: 한글 / 한자 / 영문 역명
  - 열 구성: 열차번호 | 편성 | 역1..역N | 비고
  - 미정차 = 00:00:00, 실제 시각은 15초 단위까지 존재
  - 자정 넘김: 00:03 같은 값 -> 24시 이후로 보정 (단조 증가 강제)
"""
import csv
import datetime
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/202608251a036897009480.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/home/claude/out"

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
DAY_CHAR = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}


def to_sec(t):
    """datetime.time -> 자정 기준 경과 초. 미정차(00:00:00)는 None."""
    if not isinstance(t, datetime.time):
        return None
    s = t.hour * 3600 + t.minute * 60 + t.second
    return None if s == 0 else s


def parse_days(remark):
    """비고 문자열 -> (service_id, 요일 7-플래그)"""
    r = (remark or "").strip()
    if not r or r == "매일":
        return "daily", [1] * 7
    flags = [0] * 7
    for ch in r:
        if ch in DAY_CHAR:
            flags[DAY_CHAR[ch]] = 1
    if not any(flags):
        return "daily", [1] * 7
    sid = "".join(d for d, f in zip("월화수목금토일", flags) if f)
    return sid, flags


def find_header_row(ws):
    for r in range(1, 16):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value == "열차번호":
                return r
    return None


def find_blocks(ws, hr):
    """헤더행에서 '열차번호'..'비고' 구간을 블록으로 잘라낸다."""
    starts = [c for c in range(1, ws.max_column + 1)
              if ws.cell(hr, c).value == "열차번호"]
    blocks = []
    for s in starts:
        end = None
        for c in range(s + 1, ws.max_column + 1):
            v = ws.cell(hr, c).value
            if v and str(v).startswith("비고"):
                end = c
                break
        if end:
            blocks.append((s, end))
    return blocks


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()          # name -> stop_id
    stop_meta = {}                 # name -> (hanja, eng)
    trips, stop_times, services = [], [], {}
    warnings = []

    for line_id, sheet in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet]
        hr = find_header_row(ws)
        if not hr:
            warnings.append(f"[{sheet}] 헤더행 탐지 실패 — 건너뜀")
            continue

        for b_idx, (c_start, c_remark) in enumerate(find_blocks(ws, hr)):
            title = ws.cell(hr - 1, c_start).value or ""
            direction = "up" if "상행" in str(title) else "down"
            st_cols = list(range(c_start + 2, c_remark))

            # 역 마스터 등록 (한글/한자/영문 3행)
            for c in st_cols:
                ko = ws.cell(hr, c).value
                if not ko:
                    continue
                ko = str(ko).strip()
                if ko not in stops:
                    stops[ko] = f"KR{len(stops) + 1:04d}"
                    stop_meta[ko] = (
                        str(ws.cell(hr + 1, c).value or "").strip(),
                        str(ws.cell(hr + 2, c).value or "").strip(),
                    )

            for r in range(hr + 3, ws.max_row + 1):
                train_no = ws.cell(r, c_start).value
                if train_no is None or str(train_no).strip() == "":
                    continue
                train_no = str(train_no).strip()
                formation = str(ws.cell(r, c_start + 1).value or "").strip()
                remark = ws.cell(r, c_remark).value
                sid, flags = parse_days(remark)
                services[sid] = flags

                trip_id = f"{train_no}"
                seq, prev, rolled = 0, -1, 0
                rows = []
                for c in st_cols:
                    ko = ws.cell(hr, c).value
                    if not ko:
                        continue
                    sec = to_sec(ws.cell(r, c).value)
                    if sec is None:
                        continue
                    # 자정 넘김 보정: 시각이 뒤로 가면 +24h
                    adj = sec + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = sec + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(
                            f"[{sheet}/{direction}] 열차 {train_no}: "
                            f"{ko} 시각 역행 (보정 실패)")
                    prev = adj
                    seq += 1
                    rows.append((seq, stops[str(ko).strip()], adj))

                if len(rows) < 2:
                    warnings.append(
                        f"[{sheet}/{direction}] 열차 {train_no}: 정차역 {len(rows)}개 — 확인 필요")
                    continue

                trips.append({
                    "trip_id": trip_id, "train_no": train_no,
                    "line_id": line_id, "line_name": sheet,
                    "formation": formation, "direction": direction,
                    "service_id": sid,
                    "origin": rows[0][1], "destination": rows[-1][1],
                })
                for s, sid_stop, sec in rows:
                    stop_times.append({
                        "trip_id": trip_id, "stop_seq": s,
                        "stop_id": sid_stop, "dep_sec": sec,
                    })

    import os
    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "name_hanja", "name_en"])
        for ko, sid in stops.items():
            w.writerow([sid, ko, stop_meta[ko][0], stop_meta[ko][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id", "dep_sec"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        for sid, flags in sorted(services.items()):
            w.writerow([sid] + flags)

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"calendar   {len(services):6d}")
    print(f"warnings   {len(warnings):6d}")
    for wmsg in warnings[:25]:
        print("  -", wmsg)
    if len(warnings) > 25:
        print(f"  ... 외 {len(warnings)-25}건")


if __name__ == "__main__":
    main()
