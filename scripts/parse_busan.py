#!/usr/bin/env python3
"""부산 도시철도 1/2/3/4호선(부산교통공사) 시각표 xlsx -> 정규화 테이블

원본은 4개 노선이 한 파일에 16개 시트로 들어있다(노선당 "역정보" 1개 + "평/토/휴"
시각표 3개). "역정보" 시트는 역간 거리·소요시간 참고자료일 뿐 실제 발차시각이 없어
쓰지 않는다.

시각표 시트 구조(대구·대전과 또 다름, "열차=행/역=열"로 KTX와 같은 방향):
  - 헤더 1행에 "열차번호"(4호선은 "열번") 마커가 **두 번** 나온다 — 한 시트 안에
    상행·하행 두 블록이 나란히 있는 구조(경춘선의 "광운대시종착" 시트와 같은 부류).
    블록1은 노선 기점->종점, 블록2는 그 역순.
  - 시각은 datetime 값(코레일 계열과 동일, 대구처럼 텍스트가 아님).
  - 시트명에 평/토/휴 3종 다이어 표기(대구와 같이 토요일 별도 다이어 존재).
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/부산 1~4호선 열차운행계획(2026.8.24 이후, 홈페이지용).xlsx"
# 노선 4개가 한 파일에 있지만 stops.line 값을 노선별로 나눠야 하므로(서면·동래처럼
# 노선간 실제 환승은 이름 기반 자동 병합에 맡기고) 노선당 별도 출력 폴더를 쓴다 —
# OUT_PREFIX + "1".."4" (기본: tmp/out_busan1 ~ tmp/out_busan4)
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_busan"

SHEET_RE = re.compile(r"^(\d)호선\s(평|토|휴)\(")
DAYTYPE = {"평": "평일", "토": "토요일", "휴": "휴일"}
HEADER_MARKERS = {"열차번호", "열번"}

# 다른 노선망(서울·수도권·대구·대전·국철)에 무관한 동명역이 있어 이름만 보고
# 자동 병합하면 오작동한다 — 표시명은 그대로 두고 내부 병합키만 분리(merge_key).
# DB 실제 반영 후 대조해 확정한 목록.
MERGE_KEY = {
    "중앙": "부산중앙",    # 수인분당선·안산과천선 중앙(경기 안산)과 무관
    "시청": "부산시청",    # 서울 1호선 시청과 무관
    "양정": "부산양정",    # 경의중앙선 양정(남양주)과 무관
    "금곡": "부산금곡",    # 경춘선·ITX-청춘 금곡(경기 남양주)과 무관
    "남산": "부산남산",    # 대구3호선 남산(대구 중구)과 무관
    "중동": "부산중동",    # 1호선(경인선) 중동(경기 부천)과 무관
    "증산": "부산증산",    # 서울 6호선(은평구) 증산과 무관
    "종합운동장": "부산종합운동장",  # 서울 2·9호선(송파구) 종합운동장과 무관
    # 동해선(전동)의 '부교대'→'교대' 정규화와 실제 환승역(2018년 개통 환승통로) —
    # 같은 병합키를 써서 서로만 병합되고 서울 일산선의 동명역과는 안 섞이게 한다.
    "교대": "부산교대",
}


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(name):
    return name, MERGE_KEY.get(name)


def to_sec(v):
    if not isinstance(v, (datetime.time, datetime.datetime)):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def new_line_bucket():
    return {"stops": OrderedDict(), "verify": {}, "trips": [], "stop_times": []}


def main():
    wb = load_workbook(SRC, data_only=True)
    lines = {n: new_line_bucket() for n in "1234"}
    warnings = []

    for sheet in wb.sheetnames:
        m = SHEET_RE.match(sheet)
        if not m:
            continue   # "역정보" 시트 등 — 시각표가 아니므로 건너뜀
        line_no, day_kr = m.groups()
        daytype = DAYTYPE[day_kr]
        L = lines[line_no]
        ws = wb[sheet]

        hdr = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        markers = [i for i, v in enumerate(hdr) if v in HEADER_MARKERS]

        for bi, marker_idx in enumerate(markers):
            train_col = marker_idx + 1   # 1-based openpyxl 컬럼
            end_idx = markers[bi + 1] if bi + 1 < len(markers) else len(hdr)
            station_cols = []   # (역명, 1-based 컬럼)
            for i in range(marker_idx + 1, end_idx):
                label = hdr[i]
                if not label:
                    break   # 2호선처럼 블록 길이가 짧아 뒤쪽 헤더가 빈 패딩인 경우
                name, mkey = resolve(label)
                if name not in L["stops"]:
                    L["stops"][name] = f"BS{line_no}{len(L['stops']) + 1:04d}"
                    L["verify"][name] = (label, "✔", mkey or "")
                station_cols.append((name, i + 1))

            direction = "down" if bi == 0 else "up"
            for r in range(2, ws.max_row + 1):
                train_no = clean(ws.cell(r, train_col).value)
                if not train_no:
                    continue
                raw = []
                for name, c in station_cols:
                    sec = to_sec(ws.cell(r, c).value)
                    if sec is not None:
                        raw.append([name, sec])
                if len(raw) < 2:
                    if raw:
                        warnings.append(f"[{sheet}#블록{bi}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                    continue

                rolled, prev = 0, -1
                for row in raw:
                    adj = row[1] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[1] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}#블록{bi}] {train_no}: {row[0]} 시각 역행")
                    row[1] = adj
                    prev = adj

                trip_id = f"{sheet}#{bi}#{train_no}"
                for seq, (name, sec) in enumerate(raw, start=1):
                    if seq == 1:
                        kind = "origin"
                    elif seq == len(raw):
                        kind = "destination"
                    else:
                        kind = "stop"
                    L["stop_times"].append({
                        "trip_id": trip_id, "stop_seq": seq,
                        "stop_id": L["stops"][name],
                        "arr_sec": sec, "dep_sec": sec,
                        "stop_type": kind,
                    })
                L["trips"].append({
                    "trip_id": trip_id, "train_no": train_no,
                    "line_id": f"BUSAN{line_no}", "line_name": f"부산{line_no}호선",
                    "formation": "전동차", "direction": direction,
                    "service_id": daytype,
                    "origin": L["stops"][raw[0][0]],
                    "destination": L["stops"][raw[-1][0]],
                    "next_train_no": "",
                })

    from collections import Counter
    total_stops = total_trips = total_st = 0
    for line_no, L in lines.items():
        out_dir = f"{OUT}{line_no}"
        os.makedirs(out_dir, exist_ok=True)

        with open(f"{out_dir}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["stop_id", "name_ko", "raw_label", "verify", "merge_key"])
            for name, sid in L["stops"].items():
                w.writerow([sid, name, L["verify"][name][0], L["verify"][name][1], L["verify"][name][2]])

        with open(f"{out_dir}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(L["trips"][0].keys()))
            w.writeheader()
            w.writerows(L["trips"])

        with open(f"{out_dir}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                              "arr_sec", "dep_sec", "stop_type"])
            w.writeheader()
            w.writerows(L["stop_times"])

        with open(f"{out_dir}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
            w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
            w.writerow(["토요일", 0, 0, 0, 0, 0, 1, 0])
            w.writerow(["휴일", 0, 0, 0, 0, 0, 0, 1])

        total_stops += len(L["stops"])
        total_trips += len(L["trips"])
        total_st += len(L["stop_times"])
        print(f"부산{line_no}호선  stops {len(L['stops']):4d}  trips {len(L['trips']):5d}  "
              f"stop_times {len(L['stop_times']):6d}")

    print(f"합계       stops {total_stops:4d}  trips {total_trips:5d}  stop_times {total_st:6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
