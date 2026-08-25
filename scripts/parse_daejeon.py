#!/usr/bin/env python3
"""대전 도시철도 1호선(대전교통공사) 시각표 xlsx -> 정규화 테이블

원본 구조는 다른 모든 노선과 전혀 다르다 — "열차=열" 행렬이 아니라
**출발역 기준 배차표**다:
  - 시트 = 판암행_평일 / 반석행_평일 / 판암행_휴일 / 반석행_휴일 (토요일 다이어 없음)
  - 시트 안에 역 블록이 노선 순서대로 나열되고("A역","시간" 표 헤더), 블록마다
    "시(hour) | 분(min) 분(min) ..." 행이 이어진다 — 그 역을 "출발"하는 모든 열차의
    시각 목록이다. 종착역 자체는 블록이 없다(출발할 다음 역이 없으므로).
  - 열차 단위 컬럼이 없어 "같은 열차가 역마다 어떤 시각에 나타나는가"를 직접 알 수
    없다 — 블록 간 시각 목록을 순서 보존 그리디 매칭(FIFO, 최대 간격 6분)으로 이어
    붙여 열차 단위 trip을 복원한다. 무궁화·KTX류와 달리 실제 관측 데이터(배차간격
    추정이 아니라 각 역의 실제 발차 시각 전수)라 별도 "추정" 캐비어트는 불필요하나,
    구간별로 개수가 어긋나는 지점(단축 운행 등)이 있어 그 지점의 매칭은 시각 근접도로
    복원한 것임을 README에 명시해야 한다.
  - 코레일과 마찬가지로 "역당 시각 1개"(도착/출발 구분 없음) 구조라 parse_ktx.py와
    같은 최소 스키마(stop_times.csv에 dep_sec만)를 쓴다 — build_db.py 삽입 시 KTX와
    동일한 규칙(첫 역=origin, 마지막 역=destination, 종착만 도착시각으로 해석)을 적용.
  - 대전역(도시철도)은 코레일 대전역과 지하로 직접 연결된 실제 환승역(위키백과·나무
    위키 확인) — station_groups에서 자동 병합되도록 이름을 "대전"으로 그대로 둔다.
"""
import csv
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/대전 열차운행시각표(2026.3.30.).xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_daejeon"

SHEETS = {
    "판암행_평일": ("평일", "down"), "반석행_평일": ("평일", "up"),
    "판암행_휴일": ("휴일", "down"), "반석행_휴일": ("휴일", "up"),
}
MAX_HOP_MIN = 6   # 정상 역간 간격은 1~3분 — 여유를 둔 매칭 상한
# 종착역(판암/반석) 자체는 "출발" 블록이 없어 관측 시각이 없다 — 직전 역과의
# 통상 간격(이 노선 전체에서 관측되는 역간 간격, 1~2분)만큼만 더해 도착시각을 근사한다.
TERMINUS = {"판암행_평일": "판암", "반석행_평일": "반석",
            "판암행_휴일": "판암", "반석행_휴일": "반석"}
ESTIMATED_LAST_HOP_MIN = 2

# 대전 밖 다른 노선망(대구·수도권)에 무관한 동명역이 있어 그대로 두면 build_db.py의
# 이름 기반 자동 병합이 오작동한다 — DB에 실제 반영 후 대조해 확정한 목록.
CITY_DISAMBIGUATE = {
    "용문": "용문(대전)",    # 경의중앙선 용문(양평)과 무관
    "시청": "시청(대전)",    # 서울 1호선 시청과 무관
    "중앙로": "중앙로(대전)",  # 대구1호선에도 동명역 있음
    "구암": "구암(대전)",    # 대구3호선에도 동명역 있음
}


def resolve(name):
    return CITY_DISAMBIGUATE.get(name, name)


def parse_blocks(ws):
    """[(역명, [해당 역 발차시각(분, 자정기준 누적)]), ...] — 노선 순서 그대로."""
    blocks = []
    cur_name, cur_times = None, []
    r = 1
    while r <= ws.max_row:
        a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
        if a is not None and b == "시간":
            if cur_name is not None:
                blocks.append((cur_name, sorted(cur_times)))
            cur_name, cur_times = a, []
            r += 1
            continue
        if cur_name is not None and isinstance(b, int):
            hour = b + 24 if b < 4 else b   # 자정 넘긴 새벽 시각(0시대) 보정
            for c in range(3, ws.max_column + 1):
                m = ws.cell(r, c).value
                if isinstance(m, int):
                    cur_times.append(hour * 60 + m)
        r += 1
    if cur_name is not None:
        blocks.append((cur_name, sorted(cur_times)))
    return blocks


def link_trips(blocks):
    """역 블록별 발차시각 목록을 FIFO 그리디 매칭으로 이어 붙여 열차 단위 trip 복원."""
    active = [[(0, t)] for t in blocks[0][1]]
    finished = []
    for block_idx in range(1, len(blocks)):
        times = blocks[block_idx][1]
        new_active = []
        ai = 0
        for t in times:
            matched = False
            while ai < len(active):
                last_t = active[ai][-1][1]
                if t - last_t > MAX_HOP_MIN:
                    finished.append(active[ai])
                    ai += 1
                    continue
                if t >= last_t:
                    active[ai].append((block_idx, t))
                    new_active.append(active[ai])
                    ai += 1
                    matched = True
                break
            if not matched:
                new_active.append([(block_idx, t)])
        finished.extend(active[ai:])
        active = new_active
    finished.extend(active)
    return finished


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times = [], []
    trip_seq = 0

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]
        blocks = parse_blocks(ws)
        terminus = TERMINUS[sheet]
        for name, _ in blocks:
            rname = resolve(name)
            if rname not in stops:
                stops[rname] = f"DJS{len(stops) + 1:04d}"
        if resolve(terminus) not in stops:
            stops[resolve(terminus)] = f"DJS{len(stops) + 1:04d}"

        chains = link_trips(blocks)
        for chain in chains:
            if len(chain) < 2:
                continue
            trip_seq += 1
            trip_id = f"{sheet}#{trip_seq}"
            full_run = chain[-1][0] == len(blocks) - 1
            for seq, (block_idx, t) in enumerate(chain, start=1):
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[resolve(blocks[block_idx][0])],
                    "dep_sec": t * 60,
                })
            dest_name = blocks[chain[-1][0]][0]
            if full_run:
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": len(chain) + 1,
                    "stop_id": stops[resolve(terminus)],
                    "dep_sec": (chain[-1][1] + ESTIMATED_LAST_HOP_MIN) * 60,
                })
                dest_name = terminus
            trips.append({
                "trip_id": trip_id, "train_no": f"{sheet}-{trip_seq}",
                "line_id": "DAEJEON1", "line_name": "대전1호선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops[resolve(blocks[chain[0][0]][0])],
                "destination": stops[resolve(dest_name)],
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko"])
        for name, sid in stops.items():
            w.writerow([sid, name])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id", "dep_sec"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    stop_counts = Counter(s["trip_id"] for s in stop_times)
    full = sum(1 for c in stop_counts.values() if c == 22)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}  (완주 22개 역 {full}개)")
    print(f"stop_times {len(stop_times):6d}")


if __name__ == "__main__":
    main()
