#!/usr/bin/env python3
"""대구 도시철도 1/2/3호선(대구교통공사) 시각표 xlsx -> 정규화 테이블

원본 구조(3개 파일 — 1/2/3호선 각각 별도 파일, 셋 다 같은 레이아웃):
  - 시트 = "{호선}호선({평일|토요|휴일})상행" / "...하행" (파일당 6시트)
    -> 토요일 다이어가 평일/휴일과 별도로 존재하는 첫 사례(코레일 계열엔 없었음).
  - 헤더 2행(6행=운용/편성번호로 추정, 7행=열차번호) + 8행부터 역.
    parse_line1.py류(3행 헤더)와 다르지만 "역=2행 1조(도착/출발), 열차=열" 전치구조
    자체는 동일 — parse_ilsan.py 패턴을 그대로 재사용.
  - 역명은 절삭 없이 정식 명칭. 단, 다른 호선과 동명인 역은 자기 호선 번호를 접미사로
    붙인다(예: 반월당1[1호선]/반월당2[2호선], 청라언덕2[2호선]/청라언덕3[3호선],
    명덕1[1호선]/명덕3[3호선]) — 3개 파일을 상호 대조해 확정, 접미사가 항상 자기
    파일의 호선 번호와 일치함을 확인함(나무위키 대조 불요).
  - 시트 마지막 역 몇 행 뒤에 GUID형 문자열이 낀 유령 행이 있음('-' 포함으로 필터).
  - 비코레일(대구교통공사)이라 build_db.py에서 국철과 다른 line_id로 등록해야 함.
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/대구 열차별 운행시각표(1호선) - 게시용_241007 시행.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_daegu1"

HEADER_ROW = 7          # 전 파일·전 시트 공통 확인됨(열차번호)
FIRST_STATION_ROW = 8

SHEET_RE = re.compile(r"^(\d)호선\(([^)]+)\)(상행|하행)$")
DAYTYPE = {"평일": "평일", "토요": "토요일", "휴일": "휴일"}
DIRECTION = {"상행": "up", "하행": "down"}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")


def to_sec(v):
    """이 원본은 시각이 datetime이 아니라 'HH:MM:SS' 텍스트로 들어있다."""
    if isinstance(v, (datetime.time, datetime.datetime)):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, str):
        m = TIME_RE.match(v.strip())
        if m:
            h, mi, s = (int(x) for x in m.groups())
            return h * 3600 + mi * 60 + s
    return None


def clean(v):
    return str(v).strip() if v is not None else ""


# 코레일 대구역/동대구역과 실제 도보 환승역 — "역" 표기를 벗겨 대경선·KTX와 병합되게 한다.
# "대공원"(대구2호선 연호~고산 사이)은 개통 전 가칭이었고 실제 개통명은 "수성알파시티"
# (2023년 개통, 2026-08-28 나무위키 대조로 확인)
NORMALIZE = {"대구역": "대구", "동대구역": "동대구", "대공원": "수성알파시티"}

# 대구 밖 다른 노선망에 완전히 무관한 동명역이 있어 이름만 보고 자동 병합하면
# 오작동한다(예: 대구2호선 '용산'과 서울 '용산'은 다른 역). 표시명은 그대로 두고
# 내부 병합키만 "대구"+역명으로 분리 — station_groups DB 실제 반영 후 대조해 확정한
# 목록. 노선 추가 때마다 재확인 필요.
CITY_DISAMBIGUATE = {
    "용산", "신기", "대곡", "죽전", "신천", "중앙로", "구암",
    "교대",   # 대구1호선 자체 역 — 서울 일산선·부산1호선의 동명역과 전부 무관
}


def resolve(label, line_no):
    """다른 호선과 동명인 역에 자기 호선 번호가 접미사로 붙는다 — 벗겨서 정규화.
    반환값: (정식역명, 내부 병합키(없으면 None))"""
    if label in NORMALIZE:
        return NORMALIZE[label], None
    base = label[:-1] if len(label) > 1 and label[-1] == line_no else label
    if base in CITY_DISAMBIGUATE:
        return base, f"대구{base}"
    return base, None


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []
    line_no = None

    for sheet in wb.sheetnames:
        m = SHEET_RE.match(sheet)
        if not m:
            warnings.append(f"시트명 패턴 불일치, 건너뜀: {sheet}")
            continue
        line_no, day_kr, dir_kr = m.groups()
        daytype = DAYTYPE[day_kr]
        direction = DIRECTION[dir_kr]
        ws = wb[sheet]

        st_rows = []
        for r in range(FIRST_STATION_ROW, ws.max_row + 1, 2):
            label = clean(ws.cell(r, 1).value)
            if not label or "-" in label:
                continue
            name, mkey = resolve(label, line_no)
            if name not in stops:
                stops[name] = f"DGS{len(stops) + 1:04d}"
                stop_verify[name] = (label, "✔", mkey or "")
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(HEADER_ROW, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": f"DAEGU{line_no}", "line_name": f"대구{line_no}호선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops[raw[0][0]],
                "destination": stops[raw[-1][0]],
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify", "merge_key"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1], stop_verify[name][2]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["토요일", 0, 0, 0, 0, 0, 1, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 0, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"line       대구{line_no}호선")
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
