#!/usr/bin/env python3
"""경인급행열차 시각표 xlsx -> 정규화 테이블 (parse_line1.py와 동일 스키마·같은 '1호선' 소속)

경인급행은 새 노선이 아니라 **1호선에 이미 포함된 서비스의 최신 개정판**이다.
`1호선_전동열차_*_20260228부터.xlsx`에도 `경인급행_평일_상/하` 시트가 있었지만
이 파일(2026.5.26.부터)이 더 최신이라, parse_line1.py의 SKIP_SHEETS에 옛 경인급행
시트를 추가해 빼고 이 파서로 교체했다. build_db.py에서도 별도 노선이 아니라
LINE_OF["LINE1"]("1호선")로 등록해 기존 1호선 역(구로·영등포 등)과 stop_id를 공유한다.

입력 구조 (관측):
  - 시트 = 경인급행 상행(평일) / 하행(평일) / 상행(휴일) / 하행(휴일)
  - 헤더는 경의중앙선·경춘선과 동일(3행: 시발역/종착역/열차번호, 4행부터 역)
  - '연계열번' 행 없음
  - 역명 26개 전부 이미 data/line1_stops.csv에 있는 1호선 기존 역명과 정확히
    일치(절삭 없음, 대조 완료) — 동인천~용산 구간은 원래 1호선 경인 계통 구간이라
    새로 나올 역이 없다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/경인선급행열차_20260526부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_gyeongin_express"

SHEETS = {
    "경인급행 상행(평일)": ("평일", "up"), "경인급행 하행(평일)": ("평일", "down"),
    "경인급행 상행(휴일)": ("휴일", "up"), "경인급행 하행(휴일)": ("휴일", "down"),
}

NOT_A_STATION = {"연계열번"}
JUNCTIONS = {}
STATION_ALIAS = {}  # 절삭 없음(전부 기존 1호선 역명과 일치 확인됨)

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label or label in NOT_A_STATION:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S7{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "1", "line_name": "1호선/경인급행",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
