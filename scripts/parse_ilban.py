#!/usr/bin/env python3
"""일반열차(무궁화·ITX-새마을·ITX-마음 등) 시각표 xlsx -> GTFS 유사 정규화 테이블
(parse_ktx.py와 같은 출력 스키마 — build_db.py에서 KTX와 동일하게 "국철"로 병합)

입력 구조는 KTX와 다르다 — KTX는 열차=행/역=열이지만 이 파일은 **역=행/열차=열**
(전동열차 계열과 같은 전치 구조)이면서도 역당 시각이 1개뿐이라는 점은 KTX와 같다.
시트 1개 = 노선 1개(16개), 시트 안에 하행/상행 블록이 세로로 쌓여 있다(목포보성선·
경전선·서해선처럼 하행만 있는 시트도 있음). 블록 구조(모든 시트에서 동일 확인됨):

  1~3행: 제목(한글/한자/영문), "OO선 하행" 또는 "OO선 상행"
  4~7행: 시발역 블록 — 역명/한자/영문/시각 4행 (표시용 요약, 실제 정차열 아님)
  8행:   열차종별 (무궁화/ITX-새마을/ITX-마음 등)
  9행:   열차번호
  10~N행: 역별 시각 (역명 1개 = 행 1개, 셀 값 1개 = 그 역 도착/출발 시각)
  N+1행: 비고 (요일 제한 문자, 예: "월화수목금". 비어있으면 매일)
  N+2~N+5행: 종착역 블록 (시발역 블록과 동일 구조)

행 번호를 하드코딩하지 않고 "OO행" 제목 → 열차번호 → 비고 순서로 마커를 찾아
블록 경계를 동적으로 잡는다. 미정차는 00:00:00(KTX와 동일 관례).

노선명 충돌 방지: KTX 시트와 겹치는 이름(경부선·호남선·전라선·경전선·중앙선)과
동해선(전동)·서해선(향후 반영 예정)과 겹칠 수 있는 이름 전부에 "(일반)"을 붙인다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/일반열차_20260901기준.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_ilban"

SKIP_SHEETS = {"보는방법"}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
DAY_CHAR = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}


# 부산 도시철도(1/3호선)에도 같은 이름의 무관한 역이 있어 이름만 보고 자동
# 병합하면 오작동한다(2026-08-26, 부산 노선 반영 때 발견 — 이 역은 충남 논산시
# 연산면의 호남선 연산역, 부산 연제구 연산동과 무관). 표시명은 그대로 "연산"으로
# 두고 내부 병합키만 분리한다(merge_key 컬럼).
MERGE_KEY = {
    "연산": "논산연산",
    "상동": "밀양상동",  # 서울 7호선(부천시 상동)과 무관한 경부선 상동역(밀양)
}

# 광주~광주송정 구간은 광주선 내에서 "광주송정행=상행/광주행=하행" 관례를 쓰는데
# 이 두 열차는 광주선 상행으로 진입했다가 호남선 하행으로 연계돼 방향관례가 충돌한다.
# 시트의 고정 행 순서(극락강이 광주보다 위)가 이 두 열차에서만 실제 통과 순서와
# 반대라 극락강/광주 두 항목의 위치를 swap해 바로잡는다(2026-08-30, 사용자 확인,
# rail.blue 실측 대조로 나머지 9개 역 시각은 전부 일치 확인).
GWANGJU_SPUR_SWAP_TRAINS = {"1491", "1492"}

# 이 6개 열차는 "홍성(출발)->...->평택->안중->인주->합덕->홍성(도착)"으로 돌아오는
# 서해선 관통 순환열차다. 서해선 시트에 이미 완전하고 정확한 버전이 있는 반면
# (rail.blue 실측과 초 단위까지 일치), 장항선 시트 쪽 같은 열차번호 행은 정차역
# 목록에 안중·인주·합덕이 없어 억지로 채운 값이라 신뢰 불가(예: 1246 홍성 출발시각이
# 원본 셀 자체에 18:30으로 잘못 박혀있음, 실제 15:20). 장항선 쪽은 파싱하지 않고
# 서해선 쪽 완전한 데이터만 쓴다(2026-08-30, 사용자 확인).
SEOHAE_THROUGH_TRAINS_SKIP_IN_JANGHANG = {"1241", "1242", "1243", "1244", "1245", "1246"}


def clean(v):
    return str(v).strip() if v is not None else ""


def to_sec(v):
    if not isinstance(v, (datetime.time, datetime.datetime)):
        return None
    s = v.hour * 3600 + v.minute * 60 + v.second
    return None if s == 0 else s


def parse_days(remark):
    r = clean(remark)
    if not r or r == "매일":
        return "daily", [1] * 7
    flags = [0] * 7
    for ch in r:
        if ch in DAY_CHAR:
            flags[DAY_CHAR[ch]] = 1
    if not any(flags):
        return "daily", [1] * 7
    sid = "".join(d for d, f in zip("월화수목금토일", flags) if f)
    return sid, flags


def find_blocks(ws):
    """시트 안에서 'OO선 하행/상행' 블록을 순서대로 찾는다."""
    blocks = []
    r, max_r = 1, ws.max_row
    while r <= max_r:
        title = clean(ws.cell(r, 1).value)
        if title.endswith("하행") or title.endswith("상행"):
            direction = "up" if title.endswith("상행") else "down"
            header_row = next((rr for rr in range(r, min(r + 12, max_r + 1))
                               if clean(ws.cell(rr, 1).value) == "열차번호"), None)
            if header_row is None:
                r += 1
                continue
            remark_row = next((rr for rr in range(header_row + 1, max_r + 1)
                               if clean(ws.cell(rr, 1).value).replace(" ", "") == "비고"), None)
            if remark_row is None:
                r += 1
                continue
            blocks.append({
                "title": title, "direction": direction,
                "header_row": header_row, "remark_row": remark_row,
                "station_rows": list(range(header_row + 1, remark_row)),
            })
            r = remark_row + 1
        else:
            r += 1
    return blocks


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times, services = [], [], {}
    warnings = []

    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        blocks = find_blocks(ws)
        if not blocks:
            warnings.append(f"[{sheet}] 하행/상행 블록 탐지 실패 — 건너뜀")
            continue

        for blk in blocks:
            hr, rr, direction = blk["header_row"], blk["remark_row"], blk["direction"]
            st_rows = [(r, clean(ws.cell(r, 1).value)) for r in blk["station_rows"]
                       if clean(ws.cell(r, 1).value)]
            for _, name in st_rows:
                if name not in stops:
                    stops[name] = f"KR{len(stops) + 1:04d}"

            for c in range(2, ws.max_column + 1):
                raw_train_no = ws.cell(hr, c).value
                # 실제 열차 열 뒤에 열 라벨의 한자/영문 버전이 별도 열로 붙어있다
                # (예: 열차번호 -> 列車番號 -> Train NO.). 열차번호는 항상 숫자이므로
                # 숫자가 아니면 그 열은 라벨 열로 보고 건너뛴다. 서해선 시트는 이
                # 라벨이 중간에 한 번 더 나오고 그 뒤에 열차가 더 이어지는 2블록
                # 구조라 여기서 스캔을 끝내면(break) 뒤쪽 그룹이 통째로 누락된다
                # (2026-08-30 발견, 1241/1243/1245이 서해선에서 빠져있었음) —
                # 다른 시트는 라벨 이후 유효 열이 없어 continue해도 결과가 같다.
                if raw_train_no is None or raw_train_no == "":
                    continue
                if not isinstance(raw_train_no, (int, float)):
                    continue
                train_no = clean(raw_train_no)
                if sheet == "장항선" and train_no in SEOHAE_THROUGH_TRAINS_SKIP_IN_JANGHANG:
                    continue
                formation = clean(ws.cell(hr - 1, c).value)
                sid, flags = parse_days(ws.cell(rr, c).value)
                services[sid] = flags

                rows = []
                for r, name in st_rows:
                    sec = to_sec(ws.cell(r, c).value)
                    if sec is None:
                        continue
                    rows.append((name, sec))
                if len(rows) < 2:
                    warnings.append(f"[{sheet}/{blk['title']}] {train_no}: 정차역 {len(rows)}개 — 제외")
                    continue

                if sheet == "호남선" and train_no in GWANGJU_SPUR_SWAP_TRAINS:
                    idx = {name: i for i, (name, _) in enumerate(rows)}
                    if "극락강" in idx and "광주" in idx:
                        i, j = idx["극락강"], idx["광주"]
                        rows[i], rows[j] = rows[j], rows[i]

                # ponytail: 원본에 종착역 라벨 자체가 신뢰 안 되는 행이 실제로 있었다
                # (호남선 1491/1492·장항선/서해선 1241~1246이 그런 사례였고, 위
                # GWANGJU_SPUR_SWAP_TRAINS/SEOHAE_THROUGH_TRAINS_SKIP_IN_JANGHANG로
                # 개별 확인 후 예외 처리함, 2026-08-30). 종착역 라벨을 기준으로 뒤
                # 데이터를 일괄로 잘라내는 방식은 시도했다가 다른 정상 케이스를 오히려
                # 망가뜨려서 되돌림. 그래서 자정 넘김 보정은 "큰 폭(6시간 이상)으로
                # 역행하면 다음날"로만 판단하고, 그보다 작은 역행(원본 자체의 이상치로
                # 추정)은 보정하지 않고 경고만 남긴다 — 완벽한 정합성 검증 대신 명백한
                # 24시간대 오류만 막는 선. 새 이상치가 나오면 사람이 직접 확인해 위와
                # 같은 개별 예외를 추가한다.
                ROLLBACK_THRESHOLD = 6 * 3600
                rolled, prev = 0, -1
                fixed = []
                for name, sec in rows:
                    adj = sec + rolled * 86400
                    if prev >= 0 and adj < prev and (prev - adj) >= ROLLBACK_THRESHOLD:
                        rolled += 1
                        adj = sec + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}/{blk['title']}] {train_no}: {name} 시각 역행 "
                                         f"(소폭, 원본 오류로 추정 — 보정 없이 유지)")
                    prev = adj
                    fixed.append((name, adj))

                trip_id = f"{sheet}#{blk['title']}#{train_no}"
                trips.append({
                    "trip_id": trip_id, "train_no": train_no,
                    "line_id": sheet, "line_name": f"{sheet}(일반)",
                    "formation": formation, "direction": direction,
                    "service_id": sid,
                    "origin": stops[fixed[0][0]], "destination": stops[fixed[-1][0]],
                })
                for seq, (name, adj) in enumerate(fixed, start=1):
                    stop_times.append({
                        "trip_id": trip_id, "stop_seq": seq,
                        "stop_id": stops[name], "dep_sec": adj,
                    })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "name_hanja", "name_en", "merge_key"])
        for ko, sid in stops.items():
            w.writerow([sid, ko, "", "", MERGE_KEY.get(ko, "")])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id", "dep_sec"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        for sid, flags in sorted(services.items()):
            w.writerow([sid] + flags)

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"calendar   {len(services):6d}")
    print(f"warnings   {len(warnings):6d}")
    for wmsg in warnings[:25]:
        print("  -", wmsg)
    if len(warnings) > 25:
        print(f"  ... 외 {len(warnings)-25}건")


if __name__ == "__main__":
    main()
