#!/usr/bin/env python3
"""door_directions 테이블(scripts/build_door_directions.py로 채워짐)을 사람이 보기
좋은 엑셀로 재정리한다. 역 목록은 DB 기준(미개통·유령역 제외)으로 구성하고, 순서는
원칙적으로 stops.rowid(원본 파싱 순서)를 쓰되 2026-08-28 나무위키 대조로 실제와
다른 것으로 확인된 노선(3호선·8호선)만 실제 트립의 정차 순서로 위상정렬한다.
"""
import sqlite3
import sys

import openpyxl
from openpyxl.styles import Font
from collections import defaultdict, deque

DB = sys.argv[1] if len(sys.argv) > 1 else "output/kr_rail_timetable.sqlite"
OUT = sys.argv[2] if len(sys.argv) > 2 else "output/열차_출입문_방향_260828.xlsx"

SHEET_LINE_ALIASES = {
    "3호선_일산선": ["3호선", "일산선"],
    "4호선_안산과천선": ["4호선", "안산과천선"],
    "경의중앙선": ["경의중앙선", "경의선"],
    "서해선": ["서해선(전동)"],
    "인천2호선": ["인천2호선(추정)"],
}

# rowid 순서가 실제와 다른 걸로 확인된 노선만 트립 기반 위상정렬로 교체. 나머지
# (2호선 순환선 포함)는 검증된 rowid 순서를 그대로 쓴다 — 순환선은 상행/하행이
# 방향상 반대라 위상정렬 자체가 성립하지 않는다.
TRIP_ORDERED_SHEETS = {"3호선_일산선", "8호선"}


def real_order(cur, lines):
    """실제 트립들(정차 시각 기준으로 이미 올바른 순서)에서 연속한 역 쌍을 전부
    모아 위상정렬 — 지선·분기가 있어도 트립 각각은 항상 맞는 부분순서이므로,
    합치면 노선 전체의 올바른 순서가 나온다. 상행/하행을 구분 없이 섞으면
    A->B와 B->A 간선이 동시에 생겨 사이클에 빠지므로, 하행은 뒤집어 정규화한다."""
    placeholders = ",".join("?" * len(lines))
    cur.execute(f"""
        SELECT DISTINCT t.trip_id, t.direction FROM stop_times st
        JOIN stops s ON s.stop_id = st.stop_id
        JOIN trips t ON t.trip_id = st.trip_id
        WHERE s.line IN ({placeholders})
    """, lines)
    trip_dirs = cur.fetchall()

    edges = defaultdict(set)
    nodes = {}
    for tid, direction in trip_dirs:
        cur.execute("""
            SELECT sg.group_id, sg.name_ko FROM stop_times st
            JOIN stops s ON s.stop_id = st.stop_id
            JOIN station_groups sg ON sg.group_id = s.group_id
            WHERE st.trip_id=? ORDER BY st.stop_seq
        """, (tid,))
        seq = cur.fetchall()
        if direction == "down":
            seq = seq[::-1]
        for gid, name in seq:
            nodes[gid] = name
        for i in range(len(seq) - 1):
            a, b = seq[i][0], seq[i + 1][0]
            if a != b:
                edges[a].add(b)

    indeg = defaultdict(int)
    for outs in edges.values():
        for b in outs:
            indeg[b] += 1

    q = deque(sorted(g for g in nodes if indeg[g] == 0))
    order, seen = [], set()
    indeg_local = dict(indeg)
    while q:
        g = q.popleft()
        if g in seen:
            continue
        seen.add(g)
        order.append(g)
        for b in sorted(edges.get(g, [])):
            indeg_local[b] -= 1
            if indeg_local[b] <= 0 and b not in seen:
                q.append(b)
    for g in sorted(set(nodes) - seen, key=lambda g: nodes[g]):  # 사이클 방어적 부착
        order.append(g)

    return [{"group_id": g, "name": nodes[g]} for g in order]


def rowid_order(cur, lines):
    # 실제 정차 기록(stop_times)이 0건인 유령 역(공항철도 '지상서울'·'수색',
    # 경의선 '도라산' 등 원본엔 있지만 실운행 안 하는 자리)은 제외 — 미개통역과
    # 마찬가지로 실제로 존재하지 않는 것과 같으므로 반영하지 않는다.
    placeholders = ",".join("?" * len(lines))
    cur.execute(f"""
        SELECT s.group_id, sg.name_ko, MIN(s.rowid) as ord
        FROM stops s JOIN station_groups sg ON sg.group_id = s.group_id
        WHERE s.line IN ({placeholders})
          AND EXISTS (SELECT 1 FROM stop_times st WHERE st.stop_id = s.stop_id)
        GROUP BY s.group_id ORDER BY ord
    """, lines)
    return [{"group_id": g, "name": n} for g, n, o in cur.fetchall()]


def main():
    con = sqlite3.connect(DB)
    cur = con.cursor()

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "요약"
    summary_ws.append(["노선", "역 수(DB 기준)", "문방향 확인", "부분확인/미판정", "미조사"])
    for c in summary_ws[1]:
        c.font = Font(bold=True)

    cur.execute("SELECT DISTINCT line_sheet FROM door_directions ORDER BY line_sheet")
    sheets = [r[0] for r in cur.fetchall()]

    for sheet in sheets:
        lines = SHEET_LINE_ALIASES.get(sheet, [sheet])
        db_order = real_order(cur, lines) if sheet in TRIP_ORDERED_SHEETS else rowid_order(cur, lines)
        db_gid_set = {s["group_id"] for s in db_order}
        name_by_gid = {s["group_id"]: s["name"] for s in db_order}

        # 같은 시트 안에서도 station_groups 이름 폴백으로 다른 노선 역이 잘못
        # 매칭된 경우(예: 경강선 시트에 수인분당선 역이 섞여든 것)를 걸러낸다.
        cur.execute("""
            SELECT group_id, station_name, direction, normal_side, evac_side, type, status, note, source
            FROM door_directions WHERE line_sheet=? AND group_id IS NOT NULL
        """, (sheet,))
        rows_by_gid = defaultdict(list)
        for r in cur.fetchall():
            if r[0] in db_gid_set:
                rows_by_gid[r[0]].append(r)

        ws = wb.create_sheet(title=sheet[:31])
        ws.append(["역", "방향", "평시 문방향", "대피/경합 시 문방향", "유형", "확인상태", "비고", "출처"])
        for c in ws[1]:
            c.font = Font(bold=True)

        confirmed = partial = uninvestigated = 0
        for gid in [s["group_id"] for s in db_order]:
            name = name_by_gid[gid]
            entries = rows_by_gid.get(gid)
            if not entries:
                ws.append([name, "", "", "", "", "미조사", "", ""])
                uninvestigated += 1
                continue
            for r in entries:
                ws.append([name, r[2] or "", r[3] or "", r[4] or "", r[5] or "",
                           r[6] or "", r[7] or "", r[8] or ""])
            status_text = " ".join(r[6] or "" for r in entries)
            if "확인" in status_text and "안" not in status_text and "미판정" not in status_text:
                confirmed += 1
            else:
                partial += 1

        summary_ws.append([sheet, len(db_order), confirmed, partial, uninvestigated])
        print(f"{sheet}: 역 {len(db_order)}, 확인 {confirmed}, 부분확인 {partial}, 미조사 {uninvestigated}")

    for col, width in zip("ABCDEFGH", [16, 18, 14, 18, 12, 40, 30, 50]):
        for ws in wb.worksheets:
            ws.column_dimensions[col].width = width

    wb.save(OUT)
    print("saved", OUT)


if __name__ == "__main__":
    main()
