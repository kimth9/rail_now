#!/usr/bin/env python3
"""관광열차(동해산타열차·백두대간협곡열차(V-train)·남도해양열차·서해금빛열차·정선아리랑열차)
시각표 xlsx -> 정규화 테이블 (parse_gyeongchun.py와 같은 스키마 — 역당 도착/출발 분리)

입력 구조 — 지금까지 본 어떤 노선과도 다른 비정형 레이아웃이다. 한 시트 안에 열차
6개 서비스(각 서비스마다 왕복 1~2쌍)가 작은 블록으로 흩어져 있다:

  <서비스명>(운행요일)                     <- 섹션 제목(예: "동해산타열차(목~일)")
  #2502(강릉→영주)         #2503(영주→강릉)  <- 개별 편성 제목(열차번호+구간)
  역명 | 도착 | 출발        역명 | 도착 | 출발  <- 3열 1조 헤더
  <역별 도착/출발 시각이 행마다 나열>

"역명" 셀을 전부 찾아 그 바로 위 행에서 "#번호(구간)" 제목을, 그보다 위에서 가장
가까운(같은 열 이하) 섹션 제목을 찾아 서비스명으로 묶는 방식으로 블록을 동적으로
찾는다. 시각 값이 `datetime.time`뿐 아니라 `"08:30"` 같은 문자열로도 섞여 있어
`to_sec()`이 둘 다 처리한다.

이 노선들은 전부 KTX·일반열차와 같은 국철 선로를 달리므로 stops는
LINE_OF["KTX"]("국철")로 등록해 기존 역과 공유한다.

운행요일은 "목~일" 같은 연속 구간, "목~월"처럼 주 경계를 넘는 구간, "2·7일 장날,
토~일"처럼 날짜 기반 임시 증편이 섞인 구간까지 있다. 요일 구간(~)만 calendar에
반영하고, 날짜 기반 부분(장날 등)은 flags에 반영하지 않는다 — 필요해지면
calendar_dates 방식의 예외 테이블로 확장.
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/관광열차_20260901기준.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_gwangwang"

DAY_ORDER = "월화수목금토일"
DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def clean(v):
    return str(v).strip() if v is not None else ""


def to_sec(v):
    if isinstance(v, datetime.datetime):
        v = v.time()
    if isinstance(v, datetime.time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, str):
        m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", v.strip())
        if m:
            h, mi, se = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
            return h * 3600 + mi * 60 + se
    return None


def day_flags(day_text):
    """'목~일' -> flags. 요일 구간(~)만 반영, 날짜 기반(2·7일 장날 등)은 무시."""
    m = re.search(r"([월화수목금토일])\s*~\s*([월화수목금토일])", day_text)
    flags = [0] * 7
    if not m:
        return flags
    s, e = DAY_ORDER.index(m.group(1)), DAY_ORDER.index(m.group(2))
    i = s
    while True:
        flags[i] = 1
        if i == e:
            break
        i = (i + 1) % 7
    return flags


def find_sections(ws):
    """'#'로 시작하지 않는 'OO열차(...)' 제목 셀을 전부 찾는다."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "열차(" in v and not v.startswith("#"):
                out.append((r, c, v.strip()))
    return out


def find_runs(ws):
    """'역명' 헤더 셀을 전부 찾아 (헤더행, 역명열, 편성제목) 목록을 만든다."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if clean(ws.cell(r, c).value) == "역명":
                title = clean(ws.cell(r - 1, c).value)
                out.append((r, c, title))
    return out


def section_for(sections, hr, hc):
    cands = [(r, c, t) for r, c, t in sections if r < hr]
    if not cands:
        return ""
    max_r = max(r for r, c, t in cands)
    band = [(r, c, t) for r, c, t in cands if r == max_r]
    le = [(r, c, t) for r, c, t in band if c <= hc]
    if le:
        return max(le, key=lambda x: x[1])[2]
    return min(band, key=lambda x: abs(x[1] - hc))[2]


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sections = find_sections(ws)
    runs = find_runs(ws)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times, services = [], [], {}
    warnings = []

    for hr, hc, title in runs:
        m = re.match(r"#(\d+)", title)
        if not m:
            warnings.append(f"[r{hr}c{hc}] 편성 제목에서 열차번호 추출 실패: {title!r}")
            continue
        train_no = m.group(1)
        section = section_for(sections, hr, hc)
        line_name = re.sub(r"\(.*\)$", "", section) or "관광열차"
        day_text = section

        raw = []
        r = hr + 1
        while r <= ws.max_row:
            name = clean(ws.cell(r, hc).value)
            # 블록 사이에 빈 행이 없는 경우가 있다(예: 백두대간협곡열차 V-트레인
            # 연계편 #2511 바로 다음 행에 #2514 제목이 곧바로 이어짐) — 빈 셀뿐
            # 아니라 다음 블록의 제목("#...")·헤더("역명")를 만나도 멈춘다.
            if not name or name.startswith("#") or name == "역명":
                break
            a = to_sec(ws.cell(r, hc + 1).value)
            d = to_sec(ws.cell(r, hc + 2).value)
            raw.append((name, a, d))
            r += 1
        if len(raw) < 2:
            warnings.append(f"[{title}] {train_no}: 정차역 {len(raw)}개 — 제외")
            continue

        for name, _, _ in raw:
            if name not in stops:
                stops[name] = f"SD{len(stops) + 1:04d}"
                stop_verify[name] = (name, "✔")

        sid = re.sub(r"[^가-힣0-9]", "", day_text) or "daily"
        services[sid] = day_flags(day_text)

        trip_id = f"{line_name}#{train_no}"
        for seq, (name, a, d) in enumerate(raw, start=1):
            if seq == 1:
                kind = "origin"
            elif seq == len(raw):
                kind = "destination"
            elif a is not None and d is not None:
                kind = "stop"
            else:
                kind = "pass"
            stop_times.append({
                "trip_id": trip_id, "stop_seq": seq,
                "stop_id": stops[name],
                "arr_sec": a if a is not None else "",
                "dep_sec": d if d is not None else "",
                "stop_type": kind,
            })

        trips.append({
            "trip_id": trip_id, "train_no": train_no,
            "line_id": line_name, "line_name": section or line_name,
            "formation": "관광열차", "direction": "",
            "service_id": sid,
            "origin": stops[raw[0][0]], "destination": stops[raw[-1][0]],
            "next_train_no": "",
        })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        for sid, flags in services.items():
            w.writerow([sid] + flags)

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings:
        print("  -", w_)


if __name__ == "__main__":
    main()
