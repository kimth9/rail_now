#!/usr/bin/env python3
"""부산김해경전철(비코레일, 사상~가야대) 시각표 xlsx -> 정규화 테이블

원본 구조는 부산 1~4호선과도 다르다:
  - 시트 = 평일(상선)/평일(하선)/휴일(상선)/휴일(하선) — 토요일 다이어 없음.
  - A열=역명(새 역이 시작하는 행에만), B열="도착"/"출발" 라벨(모든 데이터 행),
    C열부터 열차별 시각. 역=행 그룹(라벨로 도착/출발 명시, 종착역은 도착행만).
  - 헤더행(2행) 값이 작은 정수라 열차번호가 아니라 편성/운용번호로 보이고 하루
    안에서 반복된다 — trip_id는 컬럼 인덱스로 유일성을 보장한다.
  - 역명에 전부 "역" 접미사가 붙어있다(가야대역·사상역 등) — 다른 노선은 접미사
    없이 등록돼 있어 벗겨야 병합된다. 사상(2호선)·대저(3호선)는 실제 환승역.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/부산김해경전철_221121_열차운행시각표.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_busan_gimhae"

SHEETS = {
    "평일(상선)": ("평일", "up"), "평일(하선)": ("평일", "down"),
    "휴일(상선)": ("휴일", "up"), "휴일(하선)": ("휴일", "down"),
}


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    return label[:-1] if label.endswith("역") else label


def to_sec(v):
    if not isinstance(v, (datetime.time, datetime.datetime)):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]
        marker_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "역명")

        row_info = []   # (row, 정규화역명, '도착'|'출발')
        current = None
        for r in range(marker_row + 1, ws.max_row + 1):
            raw_label = clean(ws.cell(r, 1).value)
            if raw_label:
                current = resolve(raw_label)
                if current not in stops:
                    stops[current] = f"BG{len(stops) + 1:04d}"
                    stop_verify[current] = (raw_label, "✔")
            kind = clean(ws.cell(r, 2).value)
            if kind in ("도착", "출발") and current is not None:
                row_info.append((r, current, kind))

        for c in range(3, ws.max_column + 1):
            train_no = clean(ws.cell(2, c).value) or f"col{c}"

            raw = []
            i = 0
            while i < len(row_info):
                name = row_info[i][1]
                entry = {"name": name, "arr": None, "dep": None}
                while i < len(row_info) and row_info[i][1] == name:
                    rr, _, kind = row_info[i]
                    sec = to_sec(ws.cell(rr, c).value)
                    if kind == "도착":
                        entry["arr"] = sec
                    else:
                        entry["dep"] = sec
                    i += 1
                if entry["arr"] is not None or entry["dep"] is not None:
                    raw.append(entry)

            if len(raw) < 2:
                if raw:
                    warnings.append(f"[{sheet}] col{c}({train_no}): 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for e in raw:
                for k in ("arr", "dep"):
                    if e[k] is None:
                        continue
                    adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] col{c}({train_no}): {e['name']} 시각 역행")
                    e[k] = adj
                    prev = adj

            trip_id = f"{sheet}#{c}#{train_no}"
            for seq, e in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif e["arr"] is not None and e["dep"] is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[e["name"]],
                    "arr_sec": e["arr"] if e["arr"] is not None else "",
                    "dep_sec": e["dep"] if e["dep"] is not None else "",
                    "stop_type": kind,
                })
            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "BUSAN_GIMHAE", "line_name": "부산김해경전철",
                "formation": "경전철", "direction": direction,
                "service_id": daytype,
                "origin": stops[raw[0]["name"]],
                "destination": stops[raw[-1]["name"]],
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
