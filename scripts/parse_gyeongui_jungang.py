#!/usr/bin/env python3
"""경의중앙선 시각표 xlsx -> 정규화 테이블 (parse_line1.py와 동일 스키마)

입력 구조 (관측):
  - 2026.9.1 개정부터 경의중앙선·경의선·경춘선·서해선·ITX-청춘 5개 노선이 파일
    2개(평일/휴일)로 합본됐다(예전엔 노선별 별도 파일). 시트 이름도 요일 접미사가
    빠져 "경의중앙선 상행" / "경의중앙선 하행"으로 단순화(요일 구분은 파일 단위).
  - 전 시트가 '조정안(신규)+현행(개정전)' 대조표로 바뀌어 컬럼이 1칸 밀렸다
    (역명 2열, 열차 데이터 3열부터, 헤더도 2열 기준). 조정안 블록만 읽고 '현행'
    행부터는 무시한다(_revision_table.py).
  - '연계열번' 행이 이번 개정부터 사라짐(코레일이 원본에서 제외) -> next_train_no 공란.
  - 이 파일은 문산~지평 구간만 포함(임진강~문산 셔틀 구간은 별도 운행, 미포함)
  - 역명 특이표기 2건, 전부 나무위키·위키백과로 대조 확정:
    · '1양정'/'1양원'은 절삭이 아니라 **한국철도물류정보시스템의 내부 등록명**이다.
      위키백과 "양정역(남양주)" 문서에 "한국철도물류정보시스템에는 1양정으로
      등록되어 있다"고 명시됨 — 수인분당선의 '신인천'/'신수원'과 같은 종류의
      내부시스템 표기 함정이지만 매번 다른 노선에서 다른 역이 걸린다.
    · '도심'은 절삭이 아니라 실제 역명(도심역, 덕소~팔당 사이)이다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

from _revision_table import find_header_row, find_boundary_row, \
    collect_station_rows, find_link_row

SRC_WEEKDAY = sys.argv[1] if len(sys.argv) > 1 else \
    "RAW/korail_시각표_20260901/경의중앙선_경의선_경춘선_서해선_평일_시각표_20260901.xlsx"
SRC_HOLIDAY = sys.argv[2] if len(sys.argv) > 2 else \
    "RAW/korail_시각표_20260901/경의중앙선_경의선_경춘선_서해선_휴일_시각표_20260901.xlsx"
OUT = sys.argv[3] if len(sys.argv) > 3 else "tmp/out_gyeongui_jungang"

STATION_COL = 2

# (원본 파일, 시트, 요일, 방향)
SHEETS = [
    (SRC_WEEKDAY, "경의중앙선 상행", "평일", "up"), (SRC_WEEKDAY, "경의중앙선 하행", "평일", "down"),
    (SRC_HOLIDAY, "경의중앙선 상행", "휴일", "up"), (SRC_HOLIDAY, "경의중앙선 하행", "휴일", "down"),
]

NOT_A_STATION = {"연계열번"}
JUNCTIONS = {}  # 이 노선에는 비정차 분기점이 없음(확인됨)

# 절삭/내부시스템표기 -> 정식 역명. 전부 나무위키·위키백과 대조로 확정(미확인 항목 없음)
STATION_ALIAS = {
    "1양정": ("양정", "✔"),   # 한국철도물류정보시스템 내부 등록명(위키백과 확인)
    "1양원": ("양원", "✔"),   # 위와 동일 패턴으로 추정(양정과 동일 문서에 함께 언급된 인접역)
    "효창공": ("효창공원앞", "✔"),
    "홍대입": ("홍대입구", "✔"),
    "디엠시": ("디지털미디어시티", "✔"),
    "항공대": ("한국항공대", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wbs = {src: load_workbook(src, data_only=True) for src in {SRC_WEEKDAY, SRC_HOLIDAY}}

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for src, sheet, daytype, direction in SHEETS:
        ws = wbs[src][sheet]

        header_row = find_header_row(ws, STATION_COL)
        boundary_row = find_boundary_row(ws, header_row)

        st_rows = []
        for r, label in collect_station_rows(ws, STATION_COL, header_row, boundary_row,
                                              NOT_A_STATION, JUNCTIONS):
            if label.startswith("@"):
                st_rows.append((r, label))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S4{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        link_row = find_link_row(ws, STATION_COL, header_row, boundary_row)

        for c in range(STATION_COL + 1, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{daytype}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "GJ", "line_name": "경의중앙선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
