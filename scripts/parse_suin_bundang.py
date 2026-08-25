#!/usr/bin/env python3
"""수인분당선 시각표 xlsx -> 정규화 테이블 (parse_line1.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 평일상행 / 평일하행 / 휴일상행 / 휴일하행 (구분자 없이 붙어 있음 — 1호선과 다른 명명 규칙)
  - 전치 구조: 역 = 행, 열차 = 열 (1호선과 동일 코어)
  - 1행: 개정안(제목, 무시) / 2행: 시발역 / 3행: 종착역 / 4행: 열차번호
  - 5행부터 역이 2행 단위: (역명 행)=도착, (다음 행)=출발
  - '연계열번' 행 없음(1호선과 달리 종착 후 이어받는 열차번호 정보 없음)
  - 역명은 절삭되어 있음 -> STATION_ALIAS로 복원 (나무위키 "수도권 전철 수인·분당선"
    문서의 역 목록·환승역 표, "인천역" 문서로 전수 대조 완료)
  - '신인천'/'신수원'은 절삭이 아니라 코레일 전산상 실제 표기 — 경인선측 인천/수원역과
    선로가 분리돼 있어 별도 명칭을 쓴다(나무위키 "인천역" 문서 §2.2). 원래 역명인
    '인천'/'수원'으로 되돌린다 — build_db.py가 노선별로 승강장을 분리해 병합하므로
    1호선 쪽 인천/수원과는 station_groups에서만 합쳐지고 stops는 그대로 구분된다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/수인분당선_전동열차_20260822부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_suin_bundang"

# 시트명 -> (요일, 방향)
SHEETS = {
    "평일상행": ("평일", "up"), "평일하행": ("평일", "down"),
    "휴일상행": ("휴일", "up"), "휴일하행": ("휴일", "down"),
}

JUNCTIONS = {}  # 이 노선에는 시흥연결선 같은 비정차 분기점이 없음(확인됨)

# 절삭/전산표기 -> 정식 역명. 전부 나무위키 대조로 확정(미확인 항목 없음)
STATION_ALIAS = {
    "로데오": ("압구정로데오", "✔"),
    "강남구": ("강남구청", "✔"),
    "구룡역": ("구룡", "✔"),
    "대모산": ("대모산입구", "✔"),
    "매탄권": ("매탄권선", "✔"),
    "수원시": ("수원시청", "✔"),
    "신수원": ("수원", "✔"),   # 코레일 전산 표기(나무위키 "인천역" §2.2)
    "신인천": ("인천", "✔"),   # 코레일 전산 표기(나무위키 "인천역" §2.2)
    "신길온": ("신길온천", "✔"),
    "소래포": ("소래포구", "✔"),
    "인천논": ("인천논현", "✔"),
    "남동인": ("남동인더스파크", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        # 열차번호 헤더 행 위치 탐색 (수인분당선은 4행, 1호선은 3행 — 하드코딩하지 않는다)
        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S2{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            # 자정 넘김 보정 (단조 증가 강제)
            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"          # 급행 통과 시각
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "SB", "line_name": "수인분당선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
