#!/usr/bin/env python3
"""동해선 광역전철(부전~태화강, 부산·울산) 시각표 xlsx -> 정규화 테이블
(parse_gyeongui_jungang.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 평일 상행 / 평일 하행 / 휴일 상행 / 휴일 하행 (공백 구분)
  - 헤더는 다른 노선과 동일(3행: 시발역/종착역/열차번호, 4행부터 역). '연계열번' 있음.
  - 시종착이 '망양'인 열차가 섞여있다(부발/오이도처럼 차량기지 입출고용 중간
    시종착 — 나무위키에 "부전발 망양행 등은 차량기지 입출고로 존재"라고 명시됨).
  - 역명 절삭은 대부분 단순 3글자 프리픽스(`오시리`→오시리아, `신해운`→신해운대,
    `부산원`→부산원동, `거제해`→거제해맞이). 예외로 `부교대`→`교대`는 나무위키
    역별 승하차 순위표에 실제 표기가 '교대'로 확정됨. 단, 이 '교대'는 부산 도시철도
    1호선과 2018.2월 개통된 환승통로로 실제 연결된 환승역이라 — 서울 일산선의
    동명역 '교대'(무관)와 섞이지 않도록 표시명은 그대로 '교대'로 두고 내부
    merge_key만 `부산교대`로 등록한다(2026-08-26, 대구/대전 반영 때 build_db.py
    자동 병합이 이름만 본다는 사실이 드러나면서 발견 — 원래 있던 '교대' 그대로
    쓰면 서울 교대와 잘못 병합됨. merge_key 메커니즘 상세는 build_db.py의
    meta.station_merge_key 참조).
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/동해선_전동열차_20260228부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_donghae"

SHEETS = {
    "평일 상행": ("평일", "up"), "평일 하행": ("평일", "down"),
    "휴일 상행": ("휴일", "up"), "휴일 하행": ("휴일", "down"),
}

NOT_A_STATION = {"연계열번"}
JUNCTIONS = {}

STATION_ALIAS = {
    "오시리": ("오시리아", "✔"),
    "신해운": ("신해운대", "✔"),
    "부산원": ("부산원동", "✔"),
    "거제해": ("거제해맞이", "✔"),
    # merge_key: 표시명은 그대로 "교대"이되, 내부 병합키만 "부산교대"로 분리해
    # 서울 일산선의 동명역 '교대'(무관)와 안 섞이게 한다 — 부산1호선도 같은 키 사용.
    "부교대": ("교대", "✔", "부산교대"),
    # 서울 5호선에도 동명역(강서구 송정)이 있어 무관 — merge_key로 분리
    "송정": ("송정", "✔", "부산송정"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    """(정식역명, 검수여부, 내부 병합키(없으면 None))를 반환한다."""
    if label in STATION_ALIAS:
        v = STATION_ALIAS[label]
        return v if len(v) == 3 else (v[0], v[1], None)
    return (label, "✔", None)


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label or label in NOT_A_STATION:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf, mkey = resolve(label)
            if name not in stops:
                stops[name] = f"S9{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf, mkey or "")
            st_rows.append((r, name))

        link_row = next((r for r in range(header_row + 1, ws.max_row + 1)
                         if "연계" in clean(ws.cell(r, 1).value)), None)

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "DH", "line_name": "동해선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify", "merge_key"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1], stop_verify[name][2]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
