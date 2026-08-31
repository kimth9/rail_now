#!/usr/bin/env python3
"""안산과천선(수도권 전철 4호선 중 코레일 구간 포함 전 구간 직결) 시각표 xlsx -> 정규화 테이블
(parse_line1.py / parse_suin_bundang.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 평일(상) / 평일(하) / 휴일(상) / 휴일(하)
  - 전치 구조: 역 = 행, 열차 = 열. 헤더는 수인분당선과 동일하게 4행(열차번호)까지, 5행부터 역
  - 진접~오이도 전 구간(서울교통공사 구간 포함 51개 역)이 한 파일에 다 들어있음 —
    코레일이 직결운행 시간표를 통째로 배포하기 때문
  - '시흥기'는 정차역이 아니라 시흥기지(차량사업소) — 정거장수(나무위키 "51개") 대비
    라벨이 52개라 확인됨. JUNCTIONS로 분리
  - 역명 절삭/특수표기는 전부 나무위키 "수도권 전철 4호선" 문서로 대조 완료. 특이사항 2건:
    · '당고개'는 2024-10-31 '불암산'으로 개칭됐는데 이 원본(2026.8.22 갱신본)에 아직도
      옛 이름이 남아있다 — 코레일 내부 역명 마스터가 개칭을 못 따라간 사례
      (수인분당선의 '신인천'/'신수원'과는 다른 종류의 특이표기)
    · '4동운'은 2009년 개칭 전 '동대문운동장'의 흔적(현재는 동대문역사문화공원) — 마찬가지로
      개칭 이전 내부 코드가 남은 사례로 추정
  - '4'로 시작하는 라벨(4이촌·4서울·4충무·4동대·4창동)은 절삭이 아니라 **다른 노선에 같은
    이름의 역이 있어 붙이는 호선 구분 접두사**(4호선의 이촌/서울역/충무로/동대문/창동)
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/안산과천선_전동열차_20260822부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_ansan_gwacheon"

SHEETS = {
    "평일(상)": ("평일", "up"), "평일(하)": ("평일", "down"),
    "휴일(상)": ("휴일", "up"), "휴일(하)": ("휴일", "down"),
}

JUNCTIONS = {"시흥기": "시흥기지"}  # 승객역 아님(차량사업소) — 정거장수 51개와 대조로 확인

# 절삭/호선구분접두사/개칭전 표기 -> 정식 역명. 전부 나무위키 대조로 확정(미확인 항목 없음)
STATION_ALIAS = {
    "평촌동": ("평촌", "✔"),
    "과천청": ("정부과천청사", "✔"),
    "경마공": ("경마공원", "✔"),
    "총신대": ("총신대입구(이수)", "✔"),
    "4이촌": ("이촌", "✔"),          # 4호선 구분 접두사(경의중앙선 이촌과 구분)
    "신용산": ("신용산", "✔"),
    "숙대입": ("숙대입구", "✔"),
    "4서울": ("서울", "✔"),          # 4호선 구분 접두사. KTX/1호선과 같은 '역' 없는 표기로 맞춤
    "4충무": ("충무로", "✔"),        # 4호선 구분 접두사(3호선 충무로와 구분)
    "4동운": ("동대문역사문화공원", "✔"),  # 2009년 개칭 전 '동대문운동장' 흔적(내부 코드 미갱신)
    "4동대": ("동대문", "✔"),        # 4호선 구분 접두사(1호선 동대문과 구분)
    "한성대": ("한성대입구", "✔"),
    "성신여": ("성신여대입구", "✔"),
    "미아4": ("미아사거리", "✔"),     # '사거리'의 '사'를 숫자 4로 표기한 축약(호선 구분 접두사 아님)
    "4창동": ("창동", "✔"),          # 4호선 구분 접두사(1호선 창동과 구분)
    "당고개": ("불암산", "✔"),        # 2024-10-31 개칭. 원본 파일이 구 역명을 아직 사용 중
    "별가람": ("별내별가람", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S3{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "AG", "line_name": "안산과천선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
