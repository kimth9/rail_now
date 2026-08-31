#!/usr/bin/env python3
"""대경선(구미~경산, 대구권) 시각표 xlsx -> 정규화 테이블 (parse_gyeongui_jungang.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 평일_상 / 평일_하 / 휴일_상 / 휴일_하 (1호선과 같은 밑줄 표기)
  - 2026.9.1 개정부터 전 시트가 '조정안(신규)+현행(개정전)' 대조표로 바뀌어 컬럼이
    1칸 밀렸다(역명 2열, 열차 데이터 3열부터). 조정안 블록만 읽는다(_revision_table.py).
    '연계열번' 행도 이번 개정부터 사라짐 -> next_train_no 공란.
  - 역명 14개 전부 절삭 없이 실명 그대로다(나무위키 대조 완료) — 이 노선의 특이점은
    이름이 아니라 구성이다: 공식 정차역은 8개(경산·동대구·대구·서대구·왜관·사곡·
    북삼·구미)뿐이고, 나머지 6개(가천·고모·지천·신동·연화·약목)는 **실재하는 역이지만
    대경선 열차가 통과만 한다**(무궁화호 등 일반열차는 일부 정차). 나무위키에 "코레일
    시간표 파일에는 저 6개 역의 통과 시간도 표시되어 있다"고 명시돼 있어, 시흥기지·
    도라산과 달리 JUNCTIONS나 별도 제외 없이 일반 역으로 등록하고 기존 stop_type
    분류 로직(도착·출발 중 하나만 있으면 'pass')이 자연히 처리하게 둔다.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

from _revision_table import find_header_row, find_boundary_row, \
    collect_station_rows, find_link_row

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260901/대경선_전동열차_시각표_20260901.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_daegyeong"

STATION_COL = 2

SHEETS = {
    "평일_상": ("평일", "up"), "평일_하": ("평일", "down"),
    "휴일_상": ("휴일", "up"), "휴일_하": ("휴일", "down"),
}

NOT_A_STATION = {"연계열번"}
JUNCTIONS = {}
STATION_ALIAS = {}  # 절삭 없음(전부 실명, 나무위키 대조 완료)

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = find_header_row(ws, STATION_COL)
        boundary_row = find_boundary_row(ws, header_row)

        st_rows = []
        for r, label in collect_station_rows(ws, STATION_COL, header_row, boundary_row,
                                              NOT_A_STATION, JUNCTIONS):
            if label.startswith("@"):
                st_rows.append((r, label))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"SA{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        link_row = find_link_row(ws, STATION_COL, header_row, boundary_row)

        for c in range(STATION_COL + 1, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "DG", "line_name": "대경선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
