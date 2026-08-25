#!/usr/bin/env python3
"""경춘선 시각표 xlsx -> 정규화 테이블 (parse_gyeongui_jungang.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 경춘선 상행(평일) / 하행(평일) / 상행(휴일) / 하행(휴일) — 헤더는
    경의중앙선과 동일(3행: 시발역/종착역/열차번호, 4행부터 역), 마지막 행 '연계열번'
  - '광운대시종착(평일)' 시트가 별도로 있다. **한 시트 안에 방향이 다른 두 블록이
    좌우로 나란히 붙어있는 특이 구조**: A열=상행(춘천→광운대, 2~3열),
    E열=하행(광운대→춘천, 6~7열). 나무위키에 "춘천발 광운대행 열차는 평일에만
    하루에 2회 운행한다"고 명시돼 있어 딱 맞는 편수(4편=상행2+하행2). 휴일에는
    이 시트가 아예 없다(평일 한정 서비스).
  - 역명은 '평내호'→'평내호평' 1건만 복원하면 됨(나무위키 승하차 순위표로 확정,
    나머지는 전부 실명). 광운대는 1호선 소속 역이지만 망우선을 통해 경춘선과
    직결되는 실제 종점이라 그대로 사용.
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/경춘선_전동열차_20260228부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_gyeongchun"

# (시트, 요일, 방향, 역명 열, 데이터 열 범위)
BLOCKS = [
    ("경춘선 상행(평일)", "평일", "up", 1, None),
    ("경춘선 하행(평일)", "평일", "down", 1, None),
    ("경춘선 상행(휴일)", "휴일", "up", 1, None),
    ("경춘선 하행(휴일)", "휴일", "down", 1, None),
    ("광운대시종착(평일)", "평일", "up", 1, (2, 4)),      # 춘천 -> 광운대
    ("광운대시종착(평일)", "평일", "down", 5, (6, 8)),     # 광운대 -> 춘천
]

NOT_A_STATION = {"연계열번"}
JUNCTIONS = {}

STATION_ALIAS = {
    "평내호": ("평내호평", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, daytype, direction, station_col, col_range in BLOCKS:
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6)
                           if clean(ws.cell(r, station_col).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, station_col).value)
            if not label or label in NOT_A_STATION:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S6{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        link_row = next((r for r in range(header_row + 1, ws.max_row + 1)
                         if "연계" in clean(ws.cell(r, station_col).value)), None)

        c_start, c_end = col_range if col_range else (station_col + 1, ws.max_column)
        for c in range(c_start, c_end + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no or train_no == "열차번호":
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}/{direction}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}/{direction}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{direction}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "GC", "line_name": "경춘선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
