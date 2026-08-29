#!/usr/bin/env python3
"""1호선 시각표 xlsx -> 정규화 테이블 (KTX 파서와 동일 스키마)

입력 구조 (관측):
  - 시트 = 노선계통_요일_방향 (경부장항/서동탄-의정부/경인/경인급행/광명셔틀 x 상·하)
  - '서울급행' 시트(평일 전용, K1902/K1904/K1945)는 조정안-현행 비교표지만 두 블록이
    동일 데이터라 조정안 블록만 읽으면 됨. 다만 역=행 표가 좌(2열, 상행)/우(1열, 하행)
    블록으로 나란히 있어 다른 시트와 레이아웃이 달라 main()에서 별도 처리.
  - 전치 구조: 역 = 행, 열차 = 열
  - 2026.9.1 개정부터 전 시트가 '조정안(신규)+현행(개정전)' 대조표로 바뀌어 컬럼이
    1칸 밀렸다(역명 2열, 열차 데이터 3열부터). 헤더도 2열 기준(시발역/종착역/열차번호).
    조정안 블록만 읽고 '현행' 행부터는 무시한다(_revision_table.py).
  - B열 헤더 다음 행부터 역이 2행 단위: (역명 행)=도착, (다음 행)=출발
  - '연계열번' 행이 이번 개정부터 사라짐(코레일이 원본에서 제외) -> next_train_no 공란
  - 셀 패턴 의미:
      -D  첫 역   -> 시발 (출발만)
      AD  중간역  -> 정차
      -D  중간역  -> 통과 (통과 시각)
      A-  마지막  -> 종착 (도착만)
  - 역명은 3글자로 절삭되어 있음 -> STATION_ALIAS로 복원
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

from _revision_table import find_header_row, find_boundary_row, \
    collect_station_rows, find_link_row

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260901/1호선_평일_시각표_20260901.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_line1"

STATION_COL = 2

SKIP_SHEETS = {"서울급급행",
               # 경인급행은 별도 파일(parse_gyeongin_express.py, 2026.5.26.부터 갱신본)로
               # 대체됐다. 이 파일의 경인급행 시트는 더 오래된 데이터라 제외.
               "경인급행_평일_상", "경인급행_평일_하", "경인급행_휴일_상", "경인급행_휴일_하"}
NOT_A_STATION = {"연계열번"}

# 승강장이 아닌 통과 지점 (선로 분기점 등) — 역 마스터에서 제외
JUNCTIONS = {"시흥연": "시흥연결선"}

# 절삭된 3글자 라벨 -> 정식 역명.
# ✔ = 표기 그대로거나 자명함 / ? = 확인 필요 (verify 컬럼으로 표시)
STATION_ALIAS = {
    "1동대": ("동대문", "✔"),
    "1종로": ("종로3가", "✔"),
    "1지청": ("청량리(지하)", "✔"),
    "가산디": ("가산디지털단지", "✔"),
    "금천구": ("금천구청", "✔"),
    "동두중": ("동두천중앙", "✔"),
    "백마고": ("백마고지", "✔"),
    "성균관": ("성균관대", "✔"),
    "쌍용나": ("쌍용(나사렛대)", "✔"),

    "온양온": ("온양온천", "✔"),
    "종로5": ("종로5가", "✔"),
    "지하서": ("서울(지하)", "✔"),
    "평지제": ("평택지제", "✔"),
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, datetime.time)


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def collect_stops(ws, station_col, header_row, boundary_row, stops, stop_verify):
    """역명 열 하나를 스캔해 (행, 역명/분기점ID) 목록을 반환하고 stops를 채운다."""
    st_rows = []
    for r, label in collect_station_rows(ws, station_col, header_row, boundary_row,
                                          NOT_A_STATION, JUNCTIONS):
        if label.startswith("@"):
            st_rows.append((r, label))
            continue
        name, vf = resolve(label)
        if name not in stops:
            stops[name] = f"S1{len(stops) + 1:04d}"
            stop_verify[name] = (label, vf)
        st_rows.append((r, name))
    return st_rows


def parse_trip_column(ws, c, header_row, st_rows, stops, sheet_label, line_name,
                       direction, daytype, link_row, warnings, trips, stop_times):
    """열 하나(=열차 1편)를 읽어 trips/stop_times에 append. 유효 정차 부족 시 False."""
    train_no = clean(ws.cell(header_row, c).value)
    if not train_no:
        return False

    raw = []
    for r, name in st_rows:
        a = to_sec(ws.cell(r, c).value)
        d = to_sec(ws.cell(r + 1, c).value)
        if a is None and d is None:
            continue
        raw.append([name, a, d])
    if len(raw) < 2:
        warnings.append(f"[{sheet_label}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
        return False

    # 자정 넘김 보정 (단조 증가 강제)
    rolled, prev = 0, -1
    for row in raw:
        for i in (1, 2):
            if row[i] is None:
                continue
            adj = row[i] + rolled * 86400
            if prev >= 0 and adj < prev:
                rolled += 1
                adj = row[i] + rolled * 86400
            if prev >= 0 and adj < prev:
                warnings.append(f"[{sheet_label}] {train_no}: {row[0]} 시각 역행")
            row[i] = adj
            prev = adj

    trip_id = f"{sheet_label}#{train_no}"
    for seq, (name, a, d) in enumerate(raw, start=1):
        if seq == 1:
            kind = "origin"
        elif seq == len(raw):
            kind = "destination"
        elif a is not None and d is not None:
            kind = "stop"
        elif name.startswith("@"):
            kind = "junction"      # 선로 분기점 통과
        else:
            kind = "pass"          # 급행 통과 시각
        stop_times.append({
            "trip_id": trip_id, "stop_seq": seq,
            "stop_id": name if name.startswith("@") else stops[name],
            "arr_sec": a if a is not None else "",
            "dep_sec": d if d is not None else "",
            "stop_type": kind,
        })

    trips.append({
        "trip_id": trip_id, "train_no": train_no,
        "line_id": "1", "line_name": line_name,
        "formation": "전동차", "direction": direction,
        "service_id": daytype,
        "origin": stops.get(raw[0][0], raw[0][0]),
        "destination": stops.get(raw[-1][0], raw[-1][0]),
        "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
    })
    return True


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet in wb.sheetnames:
        if sheet == "서울급행":
            # 다른 시트와 달리 역=행 표가 좌(천안/신창→서울, 2열) · 우(서울→천안, 1열)
            # 두 블록으로 나란히 앉아 있음(각 블록의 역명 열·구분 열이 다름).
            ws = wb[sheet]
            header_row = find_header_row(ws, STATION_COL)
            boundary_row = find_boundary_row(ws, header_row)
            for station_col, train_cols, direction in (
                (2, range(3, 8), "up"),          # 천안/신창 -> 서울
                (8, range(9, ws.max_column + 1), "down"),  # 서울 -> 천안
            ):
                st_rows = collect_stops(ws, station_col, header_row, boundary_row,
                                         stops, stop_verify)
                for c in train_cols:
                    parse_trip_column(ws, c, header_row, st_rows, stops, sheet,
                                       "1호선/경부장항", direction, "평일",
                                       None, warnings, trips, stop_times)
            continue

        if sheet in SKIP_SHEETS:
            warnings.append(f"[{sheet}] 시각표 구조 아님 — 제외")
            continue
        ws = wb[sheet]
        parts = sheet.split("_")
        line_group = parts[0]
        daytype = parts[1] if len(parts) > 2 else "평일"
        direction = "up" if sheet.endswith("_상") else "down"

        header_row = find_header_row(ws, STATION_COL)
        boundary_row = find_boundary_row(ws, header_row)

        # 역 행 위치 수집 (2행 단위, 조정안 블록만)
        st_rows = collect_stops(ws, STATION_COL, header_row, boundary_row, stops, stop_verify)

        link_row = find_link_row(ws, STATION_COL, header_row, boundary_row)

        for c in range(STATION_COL + 1, ws.max_column + 1):
            parse_trip_column(ws, c, header_row, st_rows, stops, sheet,
                               f"1호선/{line_group}", direction, daytype,
                               link_row, warnings, trips, stop_times)

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
