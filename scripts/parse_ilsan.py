#!/usr/bin/env python3
"""일산선(서울 지하철 3호선 중 코레일 구간 포함 오금~대화 전 구간 직결) 시각표 xlsx
-> 정규화 테이블 (parse_ansan_gwacheon.py와 같은 패턴 — 코레일이 직결 전체 노선을 배포)

입력 구조 (관측):
  - 시트 = 일산(3호)선 평일_상행 / 평일_하행 / 휴일_상행 / 휴일_하행
  - 헤더는 다른 노선과 동일(3행: 시발역/종착역/열차번호, 4행부터 역). '연계열번' 없음.
  - 오금~대화 전 44개 역(서울교통공사 구간 포함)이 한 파일에 다 들어있다 — 안산과천선과
    같은 패턴(직결운행 전체 시각표를 코레일이 배포).
  - `3`으로 시작/끝나는 라벨(3수서·3도곡·3옥수·3충무·3을지·3종로·3대곡·금호3)은 절삭이
    아니라 **다른 노선에 같은 이름의 역이 있어 붙는 3호선 구분 표기**다(안산과천선의
    `4`접두사와 같은 부류). 예: `3수서`=수서(수인분당선·국철과 동명), `3충무`=충무로
    (안산과천선의 `4충무`와 짝을 이루는 4호선측 충무로와 환승). 나무위키 "서울 지하철
    3호선/역 목록" 문서로 전수 대조 완료, 미확인 없음.
  - **서울교통공사가 자체 배포한 "서울3호선" 파일(2023-04-03)과 실제 운행 데이터가
    거의 동일**함을 확인(2026-08-26) — 대부분 중복이라 처음엔 반영을 스킵했으나,
    사용자가 "출퇴근 증편이 있었을 텐데" 확인을 요청해 재대조한 결과 실제로 이
    파일에만 있는 열차 5개(하행 3편·상행 2편, 전부 출퇴근 시간대)를 발견 — 아래
    보충 로직으로 그 5개만 추가한다. 전체 교체를 안 하는 이유는 이 파일에 휴일
    시트가 없어서(평일만 있고 "주말동일"이라고만 적혀있음, 실제로 같은지 미검증).
    대조 시 자정 넘김 시각 표기가 두 파일에서 다르다는 것도 발견함(일산선 원본은
    0시로 랩어라운드하는 datetime, 서울3호선 원본은 "24:02:00"처럼 24시+ 텍스트) —
    절대 초가 아니라 하루 중 시각(mod 86400)으로 비교해야 이미 있는 열차가 자정
    넘김 표기 차이 때문에 "새 열차"로 잘못 잡히지 않는다.
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/일산선_20220801부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_ilsan"
# 서울교통공사가 자체 배포한 3호선 파일(2023-04-03, 일산선 원본 2022-08-01보다
# 최신) — 실제로는 같은 오금~대화 직결 운행 데이터라 대부분 중복이지만, 대조
# 결과 출퇴근시간대 열차 5개(하행 3편·상행 2편)가 이 파일에만 있었다(사용자
# 확인 요청, 2026-08-26). 그 5개만 추려서 보충한다 — 전체를 이걸로 교체하지
# 않는 이유는 이 파일에 휴일 시트가 없어서(평일만 있고 "주말동일"이라고만
# 적혀있음, 실제로 같은지는 미검증) 기존 일산선 휴일 데이터를 그대로 신뢰하는
# 편이 안전하기 때문.
SRC_SEOUL3 = "RAW/230403_서울3호선 평일증편, 주말동일.xlsx"

SHEETS = {
    "일산(3호)선 평일_상행": ("평일", "up"), "일산(3호)선 평일_하행": ("평일", "down"),
    "일산(3호)선 휴일_상행": ("휴일", "up"), "일산(3호)선 휴일_하행": ("휴일", "down"),
}

NOT_A_STATION = set()
JUNCTIONS = {}

STATION_ALIAS = {
    "경찰병": ("경찰병원", "✔"),
    "가락시": ("가락시장", "✔"),
    "3수서": ("수서", "✔"),         # 3호선 구분 표기(수인분당선·국철 수서와 동명)
    "3도곡": ("도곡", "✔"),         # 3호선 구분 표기(수인분당선 도곡과 동명)
    "남부터": ("남부터미널", "✔"),
    "고속터": ("고속터미널", "✔"),
    "3옥수": ("옥수", "✔"),         # 3호선 구분 표기(경의중앙선 옥수와 동명)
    "금호3": ("금호", "✔"),
    "동대입": ("동대입구", "✔"),     # 정식 역명 자체가 '동대입구'(동대문입구 아님)
    "3충무": ("충무로", "✔"),       # 3호선 구분 표기(안산과천선 4충무=충무로와 환승)
    "3을지": ("을지로3가", "✔"),
    "3종로": ("종로3가", "✔"),      # 3호선 구분 표기(1호선 종로3가와 동명)
    "3대곡": ("대곡", "✔"),         # 3호선 구분 표기(경의중앙선·경의선 대곡과 동명)
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")


def to_sec(v):
    if is_time(v):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, str):
        m = TIME_RE.match(v.strip())
        if m:
            h, mi, s = (int(x) for x in m.groups())
            return h * 3600 + mi * 60 + s
    return None


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label or label in NOT_A_STATION:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"SB{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "IS", "line_name": "일산선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": "",
            })

    # --- 서울3호선 파일에서 일산선에 없는 출퇴근 증편 열차만 보충 ---
    # 기종점(origin/destination)만으로는 하루에 같은 구간을 오가는 열차가 수십
    # 편이라 구분이 안 된다 — 첫 역 출발시각(분 단위, 자정 기준 초)까지 같이
    # 봐야 "이미 있는 열차"와 "진짜 새 열차"가 구분된다.
    stop_times_by_trip = {}
    for st in stop_times:
        stop_times_by_trip.setdefault(st["trip_id"], []).append(st)
    existing_sig = set()
    for t in trips:
        if t["service_id"] != "평일":
            continue
        first = min(stop_times_by_trip[t["trip_id"]], key=lambda s: int(s["stop_seq"]))
        fd = first["dep_sec"] or first["arr_sec"]
        # 두 원본이 자정 넘김 표기가 다르다(일산선=0시로 랩어라운드, 서울3호선=24시+
        # 표기) — 절대 초가 아니라 하루 중 시각(mod 86400)으로 비교해야 같은 실제
        # 열차가 다르게 안 잡힌다.
        existing_sig.add((t["direction"], t["origin"], t["destination"], (int(fd) % 86400) // 60))
    wb3 = load_workbook(SRC_SEOUL3, data_only=True)
    added = 0
    for sheet, direction in {"평일 하행": "down", "평일 상행": "up"}.items():
        ws3 = wb3[sheet]
        st_rows3 = []
        for r in range(5, ws3.max_row + 1, 2):
            label = clean(ws3.cell(r, 1).value)
            if not label or label == "연결열번":
                continue
            st_rows3.append((r, label))
        for c in range(2, ws3.max_column + 1):
            train_no = clean(ws3.cell(1, c).value)
            if not train_no:
                continue
            raw = []
            for r, name in st_rows3:
                a = to_sec(ws3.cell(r, c).value)
                d = to_sec(ws3.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                continue
            origin, dest = raw[0][0], raw[-1][0]
            if origin not in stops or dest not in stops:
                continue   # 역명이 다르면(있을 리 없지만) 안전하게 건너뜀
            first_dep = raw[0][2] if raw[0][2] is not None else raw[0][1]
            sig = (direction, stops[origin], stops[dest], (int(first_dep) % 86400) // 60)
            if sig in existing_sig:
                continue   # 이미 일산선에 같은 열차 있음(분 단위 출발시각까지 일치)

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    row[i] = adj
                    prev = adj

            trip_id = f"서울3호선증편#{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })
            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "IS", "line_name": "일산선",
                "formation": "전동차", "direction": direction,
                "service_id": "평일",
                "origin": stops[origin], "destination": stops[dest],
                "next_train_no": "",
            })
            added += 1
    print(f"서울3호선 보충열차 {added:6d}")

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
