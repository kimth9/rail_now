#!/usr/bin/env python3
"""공항철도(AREX, 비코레일) 시각표 xlsx -> 정규화 테이블

시트 = 평일/휴일. 각 시트 안에 하행(서울⇒인천공항)·상행(인천공항⇒서울) 두
블록이 세로로 쌓여있다(헤더가 "열차번호" 행마다 새로 나타남 — 그 행을 전부
찾아 블록 경계로 삼는다). 헤더 5행(열차번호/시발역/시발시간/종착역/종착시간)
+ 역=2행1조(도착/출발) 전치 구조, 텍스트 시각("05:20:00").

- `수색직결선`은 실제 정차역이 아니라 선로 연결 지점(값이 `---`로 표시돼
  통과만 함) — JUNCTIONS로 분리.
- `지상서울`은 원본에 자리는 있지만 두 시트·양방향 전부 실제 정차 기록이
  0건(코레일 본선 지상 연결용 예비 표기로 추정) — 등록은 하되 "미사용 역"으로
  진단 출력.
"""
import csv
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/251229_공항철도_열차시각표_251229.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_airport_railroad"

JUNCTIONS = {"수색직결선": "수색직결선"}
TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")


def clean(v):
    return str(v).strip() if v is not None else ""


def to_sec(v):
    if not isinstance(v, str):
        return None
    m = TIME_RE.match(v.strip())
    if not m:
        return None
    h, mi, s = (int(x) for x in m.groups())
    return h * 3600 + mi * 60 + s


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times = [], []
    warnings = []

    for sheet in wb.sheetnames:
        daytype = sheet
        ws = wb[sheet]
        header_rows = [r for r in range(1, ws.max_row + 1)
                       if clean(ws.cell(r, 1).value) == "열차번호"]

        for bi, header_row in enumerate(header_rows):
            # 다음 블록의 헤더 "직전 행"은 그 블록의 제목(하행/상행)이라 역이 아님
            block_end = (header_rows[bi + 1] - 1) if bi + 1 < len(header_rows) else ws.max_row + 1
            title = clean(ws.cell(header_row - 1, 1).value) + clean(ws.cell(header_row - 1, 2).value)
            direction = "down" if "⇒인천공항" in title.replace(" ", "") else "up"

            first_station_row = header_row + 5   # 시발역/시발시간/종착역/종착시간 4행 건너뜀
            st_rows = []
            for r in range(first_station_row, block_end):
                label = clean(ws.cell(r, 1).value)
                if not label or label.replace(" ", "") == "비고":
                    continue
                if label in JUNCTIONS:
                    st_rows.append((r, "@" + JUNCTIONS[label]))
                    continue
                if label not in stops:
                    stops[label] = f"AR{len(stops) + 1:04d}"
                st_rows.append((r, label))

            for c in range(2, ws.max_column + 1):
                train_no = clean(ws.cell(header_row, c).value)
                if not train_no:
                    continue
                raw = []
                for r, name in st_rows:
                    a = to_sec(ws.cell(r, c).value)
                    d = to_sec(ws.cell(r + 1, c).value)
                    if a is None and d is None:
                        continue
                    raw.append([name, a, d])
                if len(raw) < 2:
                    if raw:
                        warnings.append(f"[{sheet}#{bi}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                    continue

                rolled, prev = 0, -1
                for row in raw:
                    for i in (1, 2):
                        if row[i] is None:
                            continue
                        adj = row[i] + rolled * 86400
                        if prev >= 0 and adj < prev:
                            rolled += 1
                            adj = row[i] + rolled * 86400
                        if prev >= 0 and adj < prev:
                            warnings.append(f"[{sheet}#{bi}] {train_no}: {row[0]} 시각 역행")
                        row[i] = adj
                        prev = adj

                trip_id = f"{daytype}#{bi}#{train_no}"
                for seq, (name, a, d) in enumerate(raw, start=1):
                    if seq == 1:
                        kind = "origin"
                    elif seq == len(raw):
                        kind = "destination"
                    elif a is not None and d is not None:
                        kind = "stop"
                    elif name.startswith("@"):
                        kind = "junction"
                    else:
                        kind = "pass"
                    stop_times.append({
                        "trip_id": trip_id, "stop_seq": seq,
                        "stop_id": name if name.startswith("@") else stops[name],
                        "arr_sec": a if a is not None else "",
                        "dep_sec": d if d is not None else "",
                        "stop_type": kind,
                    })
                trips.append({
                    "trip_id": trip_id, "train_no": train_no,
                    "line_id": "AREX", "line_name": "공항철도",
                    "formation": "전동차", "direction": direction,
                    "service_id": daytype,
                    "origin": stops.get(raw[0][0], raw[0][0]),
                    "destination": stops.get(raw[-1][0], raw[-1][0]),
                    "next_train_no": "",
                })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko"])
        for name, sid in stops.items():
            w.writerow([sid, name])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)

    used = {s["stop_id"] for s in stop_times}
    unused = [name for name, sid in stops.items() if sid not in used]
    if unused:
        print(f"미사용 역(정차 기록 0건): {unused}")


if __name__ == "__main__":
    main()
