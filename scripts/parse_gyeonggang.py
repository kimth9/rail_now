#!/usr/bin/env python3
"""경강선(판교~여주) 시각표 xlsx -> 정규화 테이블 (parse_gyeongui_jungang.py와 동일 스키마)

입력 구조 (관측):
  - 시트 = 평일 상행 / 평일 하행 / 휴일 상행 / 휴일 하행 (공백 구분, 방향이 뒤에 옴)
  - 헤더는 경의중앙선과 동일(3행: 시발역/종착역/열차번호, 4행부터 역). '연계열번' 없음.
  - 역명 특이표기: `신이매`/`신판교`는 절삭이 아니라 **수인분당선의 이매·신분당선의
    판교와 같은 이름이라 코레일 내부 시스템이 구분하려고 붙인 접두사**다(수인분당선의
    `신인천`/`신수원`과 동일한 부류의 함정). 나무위키 "경강선/환승역" 표에 각각
    "이매(수인·분당선 환승)", "판교(신분당선 환승)"로 단독 표기돼 있어 확인됨 —
    정식 역명으로 되돌리면 `build_db.py`에서 수인분당선 이매와 station_groups로
    정상 병합된다(판교는 신분당선이 DB에 없어 단독 역으로 남음).
  - `세종릉`→`세종대왕릉`, `도예촌`→`신둔도예촌`, `경광주`→`경기광주`는 단순 3글자
    절삭이 아니라 중간 음절을 건너뛴 압축 표기다. 전부 나무위키 역별 승하차 순위표로
    대조 완료(미확인 항목 없음).
"""
import csv
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/korail_시각표_20260825/경강선_전동열차_20260323부터.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_gyeonggang"

SHEETS = {
    "평일 상행": ("평일", "up"), "평일 하행": ("평일", "down"),
    "휴일 상행": ("휴일", "up"), "휴일 하행": ("휴일", "down"),
}

NOT_A_STATION = set()
JUNCTIONS = {}

STATION_ALIAS = {
    "세종릉": ("세종대왕릉", "✔"),
    "도예촌": ("신둔도예촌", "✔"),
    "경광주": ("경기광주", "✔"),
    "신이매": ("이매", "✔"),   # 코레일 내부 구분 표기(수인분당선 이매와 동명)
    "신판교": ("판교", "✔"),   # 코레일 내부 구분 표기(신분당선 판교와 동명)
}

DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def is_time(v):
    return isinstance(v, (datetime.time, datetime.datetime))


def to_sec(v):
    if not is_time(v):
        return None
    return v.hour * 3600 + v.minute * 60 + v.second


def clean(v):
    return str(v).strip() if v is not None else ""


def resolve(label):
    if label in STATION_ALIAS:
        return STATION_ALIAS[label]
    return (label, "✔")


def main():
    wb = load_workbook(SRC, data_only=True)

    stops, stop_verify = OrderedDict(), {}
    trips, stop_times = [], []
    warnings = []

    for sheet, (daytype, direction) in SHEETS.items():
        ws = wb[sheet]

        header_row = next(r for r in range(1, 6) if clean(ws.cell(r, 1).value) == "열차번호")

        st_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = clean(ws.cell(r, 1).value)
            if not label or label in NOT_A_STATION:
                continue
            if label in JUNCTIONS:
                st_rows.append((r, "@" + JUNCTIONS[label]))
                continue
            name, vf = resolve(label)
            if name not in stops:
                stops[name] = f"S8{len(stops) + 1:04d}"
                stop_verify[name] = (label, vf)
            st_rows.append((r, name))

        for c in range(2, ws.max_column + 1):
            train_no = clean(ws.cell(header_row, c).value)
            if not train_no:
                continue

            raw = []
            for r, name in st_rows:
                a = to_sec(ws.cell(r, c).value)
                d = to_sec(ws.cell(r + 1, c).value)
                if a is None and d is None:
                    continue
                raw.append([name, a, d])
            if len(raw) < 2:
                warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for row in raw:
                for i in (1, 2):
                    if row[i] is None:
                        continue
                    adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = row[i] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {row[0]} 시각 역행")
                    row[i] = adj
                    prev = adj

            trip_id = f"{sheet}#{train_no}"
            for seq, (name, a, d) in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif a is not None and d is not None:
                    kind = "stop"
                elif name.startswith("@"):
                    kind = "junction"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": name if name.startswith("@") else stops[name],
                    "arr_sec": a if a is not None else "",
                    "dep_sec": d if d is not None else "",
                    "stop_type": kind,
                })

            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "GG", "line_name": "경강선",
                "formation": "전동차", "direction": direction,
                "service_id": daytype,
                "origin": stops.get(raw[0][0], raw[0][0]),
                "destination": stops.get(raw[-1][0], raw[-1][0]),
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko", "raw_label", "verify"])
        for name, sid in stops.items():
            w.writerow([sid, name, stop_verify[name][0], stop_verify[name][1]])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id"] + DAYS)
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    from collections import Counter
    kinds = Counter(s["stop_type"] for s in stop_times)
    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}  {dict(kinds)}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
