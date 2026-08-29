#!/usr/bin/env python3
"""output/열차_출입문_방향_*.xlsx를 읽어 기존 SQLite DB에 door_directions 테이블로
얹는다. build_db.py가 DB를 통째로 재생성하므로 그 뒤에 실행해야 하며, 엑셀을
계속 채워나갈 때마다 이 스크립트만 재실행하면 된다(door_directions만 교체).
"""
import sqlite3
import sys

import openpyxl

DB = sys.argv[1] if len(sys.argv) > 1 else "output/kr_rail_timetable.sqlite"
XLSX = sys.argv[2] if len(sys.argv) > 2 else "output/열차_출입문_방향_260828.xlsx"

# 시트명이 stops.line 값과 다르거나(예: 서해선(전동)) 두 노선을 합쳐 둔 시트만 명시.
# 나머지는 시트명을 stops.line 값으로 그대로 사용한다.
SHEET_LINE_ALIASES = {
    "3호선_일산선": ["3호선", "일산선"],
    "4호선_안산과천선": ["4호선", "안산과천선"],
    "경의중앙선": ["경의중앙선", "경의선"],
    "서해선": ["서해선(전동)"],
    "인천2호선": ["인천2호선(추정)"],
}

SCHEMA = """
DROP TABLE IF EXISTS door_directions;
CREATE TABLE door_directions(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    line_sheet TEXT, group_id TEXT, station_name TEXT, direction TEXT,
    normal_side TEXT, evac_side TEXT, type TEXT, status TEXT,
    note TEXT, source TEXT
);
"""


def main():
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.executescript(SCHEMA)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    total = matched = 0
    unmatched = []

    for sheet in wb.sheetnames:
        if sheet == "요약":
            continue
        lines = SHEET_LINE_ALIASES.get(sheet, [sheet])
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            station = row[0]
            if not station:
                continue
            total += 1
            placeholders = ",".join("?" * len(lines))
            cur.execute(
                f"SELECT DISTINCT group_id FROM stops WHERE name_ko=? AND line IN ({placeholders})",
                (station, *lines),
            )
            gids = [r[0] for r in cur.fetchall()]
            if len(gids) != 1:
                cur.execute("SELECT DISTINCT group_id FROM station_groups WHERE name_ko=?", (station,))
                gids = [r[0] for r in cur.fetchall()]
            group_id = gids[0] if len(gids) == 1 else None
            if group_id:
                matched += 1
            else:
                unmatched.append((sheet, station))
            cur.execute(
                "INSERT INTO door_directions VALUES(NULL,?,?,?,?,?,?,?,?,?,?)",
                (sheet, group_id, station, *row[1:8]),
            )

    con.commit()
    print(f"{total}행 중 {matched}행 group_id 매칭, {total - matched}행 미매칭")
    if unmatched:
        print("미매칭 목록:")
        for sheet, station in unmatched:
            print(f"  {sheet}: {station}")


if __name__ == "__main__":
    main()
