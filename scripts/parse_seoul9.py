#!/usr/bin/env python3
"""서울지하철 9호선(서울교통공사) 시각표 xlsx -> 정규화 테이블

평일/휴일이 완전히 별개 파일로 배포된다(9호선만 그런 게 아니라 서울교통공사
계열 전체의 특징 — 시트명이 평일·휴일 파일에서 동일해서 daytype을 파일명이
아니라 인자로 받는다). 시트 = 상행/하행 × 일반열차/급행열차 (+평일 파일에만
회송열차 시트가 있으나 비영업열차라 건너뜀).

구조는 부산김해경전철과 같은 "명시적 라벨" 방식:
  A열=No.(정차 순번, 미사용) / B열=역명(새 역 시작 행에만) / C열="도착"/"출발"
  / D열부터 열차별 시각(텍스트 'HH:MM:SS'). 급행열차는 통과역이 그냥 빈칸으로
  남아있어(행 자체는 유지) 파싱 로직상 자동으로 통과 처리된다.
"""
import csv
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/240330_붙임1. 평일 열차별운행시각표(9호선).xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_seoul9"
DAYTYPE = sys.argv[3] if len(sys.argv) > 3 else "평일"

SHEETS = {
    "상행(일반열차)": "up", "하행(일반열차)": "down",
    "상행(급행열차)": "up", "하행(급행열차)": "down",
}

TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")


def clean(v):
    return str(v).strip() if v is not None else ""


def to_sec(v):
    if isinstance(v, str):
        m = TIME_RE.match(v.strip())
        if m:
            h, mi, s = (int(x) for x in m.groups())
            return h * 3600 + mi * 60 + s
    return None


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times = [], []
    warnings = []

    for sheet, direction in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        row_info = []   # (row, 역명, '도착'|'출발')
        current = None
        for r in range(2, ws.max_row + 1):
            label = clean(ws.cell(r, 2).value)
            if label:
                current = label
                if current not in stops:
                    stops[current] = f"S9M{len(stops) + 1:04d}"
            kind = clean(ws.cell(r, 3).value)
            if kind in ("도착", "출발") and current is not None:
                row_info.append((r, current, kind))

        for c in range(4, ws.max_column + 1):
            train_no = clean(ws.cell(1, c).value)
            if not train_no:
                continue

            raw = []
            i = 0
            while i < len(row_info):
                name = row_info[i][1]
                entry = {"name": name, "arr": None, "dep": None}
                while i < len(row_info) and row_info[i][1] == name:
                    rr, _, kind = row_info[i]
                    sec = to_sec(ws.cell(rr, c).value)
                    if kind == "도착":
                        entry["arr"] = sec
                    else:
                        entry["dep"] = sec
                    i += 1
                if entry["arr"] is not None or entry["dep"] is not None:
                    raw.append(entry)

            if len(raw) < 2:
                if raw:
                    warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for e in raw:
                for k in ("arr", "dep"):
                    if e[k] is None:
                        continue
                    adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {e['name']} 시각 역행")
                    e[k] = adj
                    prev = adj

            trip_id = f"{DAYTYPE}#{sheet}#{train_no}"
            for seq, e in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif e["arr"] is not None and e["dep"] is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[e["name"]],
                    "arr_sec": e["arr"] if e["arr"] is not None else "",
                    "dep_sec": e["dep"] if e["dep"] is not None else "",
                    "stop_type": kind,
                })
            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "SEOUL9", "line_name": "9호선",
                "formation": "전동차", "direction": direction,
                "service_id": DAYTYPE,
                "origin": stops[raw[0]["name"]],
                "destination": stops[raw[-1]["name"]],
                "next_train_no": "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko"])
        for name, sid in stops.items():
            w.writerow([sid, name])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
