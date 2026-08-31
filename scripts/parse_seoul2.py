#!/usr/bin/env python3
"""서울지하철 2호선(서울교통공사) 시각표 xlsx -> 정규화 테이블

순환선이라 방향이 상행/하행이 아니라 외선(outer)/내선(inner)이다. 본선(성수~성수
순환) 43개 역만 포함 — 성수지선(신설동)·신정지선(까치산) 지선은 이 파일에 없음.

평일/휴일 원본의 **레이아웃 자체가 다르다**:
  - 평일(260821, 최신, 7호선과 합본): 헤더에 "출발역/도착역/운행구분" +
    명시적 "도착"/"출발" 라벨 열(5~8호선류와 같은 방식)
  - 휴일(220401, 2022년 원본, 4년 이상 오래됨 — 재확보 전까지 이걸로 진행):
    라벨 열 없이 암묵적 2행1조(도착/출발) 구조(1호선류와 같은 방식)
두 방식을 자동 감지해서 처리한다(header_row 다음 데이터 행의 두 번째 열 값이
"출발"/"도착"이면 명시적, 아니면 암묵적).
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "RAW/260821_서울2호선(평일)_ 서울7호선(평일) 열차운행시각표.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "tmp/out_seoul2"
DAYTYPE = sys.argv[3] if len(sys.argv) > 3 else "평일"

SHEET_DIRECTION = {"외선": "outer", "내선": "inner"}
TIME_RE = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")
NOT_A_STATION = {"연결열차", "연결열번", "운행구분"}


def clean(v):
    return str(v).strip().replace(" ", "") if v is not None else ""


def to_sec(v):
    """이 노선은 파일마다 시각 표현이 다르다 — 평일본(260821)은 datetime,
    휴일본(220401)은 관측 결과에 따라 문자열일 수 있어 둘 다 받는다."""
    if isinstance(v, (datetime.time, datetime.datetime)):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, str):
        m = TIME_RE.match(v.strip())
        if m:
            h, mi, s = (int(x) for x in m.groups())
            return h * 3600 + mi * 60 + s
    return None


def find_header(ws):
    header_row = next(r for r in range(1, 8) if clean(ws.cell(r, 1).value) == "열차번호")
    meta_labels = {"출발역", "도착역"} | NOT_A_STATION
    link_row = None
    r = header_row + 1
    while True:
        v = clean(ws.cell(r, 1).value)
        if v == "" or v in meta_labels:
            if v in ("연결열차", "연결열번"):
                link_row = r
            r += 1
            continue
        break
    return header_row, r, link_row


def parse_explicit(ws, header_row, first_station_row, stops):
    """5/6/8호선류 — B열에 "도착"/"출발" 라벨이 명시된 구조."""
    row_info = []
    current = None
    for rr in range(first_station_row, ws.max_row + 1):
        label = clean(ws.cell(rr, 1).value)
        if label:
            if label in NOT_A_STATION:
                current = None
                continue
            current = label
            if current not in stops:
                stops[current] = f"S2M{len(stops) + 1:04d}"
        kind = clean(ws.cell(rr, 2).value)
        if kind in ("도착", "출발") and current is not None:
            row_info.append((rr, current, kind))

    per_train = {}
    for c in range(3, ws.max_column + 1):
        train_no = clean(ws.cell(header_row, c).value)
        if not train_no:
            continue
        raw = []
        i = 0
        while i < len(row_info):
            name = row_info[i][1]
            entry = {"name": name, "arr": None, "dep": None}
            while i < len(row_info) and row_info[i][1] == name:
                rr, _, kind = row_info[i]
                sec = to_sec(ws.cell(rr, c).value)
                if kind == "도착":
                    entry["arr"] = sec
                else:
                    entry["dep"] = sec
                i += 1
            if entry["arr"] is not None or entry["dep"] is not None:
                raw.append(entry)
        per_train[(train_no, c)] = raw
    return per_train


def parse_implicit(ws, header_row, first_station_row, stops):
    """1호선류 — 라벨 없이 역=2행1조(도착/출발) 위치로 판별하는 구조."""
    st_rows = []
    for r in range(first_station_row, ws.max_row + 1, 2):
        label = clean(ws.cell(r, 1).value)
        if not label or label in NOT_A_STATION:
            continue
        if label not in stops:
            stops[label] = f"S2M{len(stops) + 1:04d}"
        st_rows.append((r, label))

    per_train = {}
    for c in range(2, ws.max_column + 1):
        train_no = clean(ws.cell(header_row, c).value)
        if not train_no:
            continue
        raw = []
        for r, name in st_rows:
            a = to_sec(ws.cell(r, c).value)
            d = to_sec(ws.cell(r + 1, c).value)
            if a is None and d is None:
                continue
            raw.append({"name": name, "arr": a, "dep": d})
        per_train[(train_no, c)] = raw
    return per_train


def main():
    wb = load_workbook(SRC, data_only=True)

    stops = OrderedDict()
    trips, stop_times = [], []
    warnings = []

    for sheet in wb.sheetnames:
        dir_kr = next((k for k in SHEET_DIRECTION if k in sheet), None)
        if dir_kr is None:
            continue
        # 원본 파일 하나에 평일·휴일 시트가 같이 들어있는 경우(예: 220401 구버전
        # 파일)가 있어, 시트명에 명시된 요일 종류가 있으면 요청한 DAYTYPE과
        # 다를 때 건너뛴다.
        if ("평일" in sheet and DAYTYPE != "평일") or ("휴일" in sheet and DAYTYPE != "휴일"):
            continue
        direction = SHEET_DIRECTION[dir_kr]
        ws = wb[sheet]

        header_row, first_station_row, link_row = find_header(ws)

        probe = clean(ws.cell(first_station_row, 2).value)
        if probe in ("도착", "출발"):
            per_train = parse_explicit(ws, header_row, first_station_row, stops)
        else:
            per_train = parse_implicit(ws, header_row, first_station_row, stops)

        for (train_no, c), raw in per_train.items():
            if len(raw) < 2:
                if raw:
                    warnings.append(f"[{sheet}] {train_no}: 유효 정차 {len(raw)}개 — 제외")
                continue

            rolled, prev = 0, -1
            for e in raw:
                for k in ("arr", "dep"):
                    if e[k] is None:
                        continue
                    adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        rolled += 1
                        adj = e[k] + rolled * 86400
                    if prev >= 0 and adj < prev:
                        warnings.append(f"[{sheet}] {train_no}: {e['name']} 시각 역행")
                    e[k] = adj
                    prev = adj

            trip_id = f"{DAYTYPE}#{sheet}#{train_no}#{c}"
            for seq, e in enumerate(raw, start=1):
                if seq == 1:
                    kind = "origin"
                elif seq == len(raw):
                    kind = "destination"
                elif e["arr"] is not None and e["dep"] is not None:
                    kind = "stop"
                else:
                    kind = "pass"
                stop_times.append({
                    "trip_id": trip_id, "stop_seq": seq,
                    "stop_id": stops[e["name"]],
                    "arr_sec": e["arr"] if e["arr"] is not None else "",
                    "dep_sec": e["dep"] if e["dep"] is not None else "",
                    "stop_type": kind,
                })
            trips.append({
                "trip_id": trip_id, "train_no": train_no,
                "line_id": "SEOUL2", "line_name": "2호선",
                "formation": "전동차", "direction": direction,
                "service_id": DAYTYPE,
                "origin": stops[raw[0]["name"]],
                "destination": stops[raw[-1]["name"]],
                "next_train_no": clean(ws.cell(link_row, c).value) if link_row else "",
            })

    os.makedirs(OUT, exist_ok=True)

    with open(f"{OUT}/stops.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["stop_id", "name_ko"])
        for name, sid in stops.items():
            w.writerow([sid, name])

    with open(f"{OUT}/trips.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(trips[0].keys()))
        w.writeheader()
        w.writerows(trips)

    with open(f"{OUT}/stop_times.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["trip_id", "stop_seq", "stop_id",
                                          "arr_sec", "dep_sec", "stop_type"])
        w.writeheader()
        w.writerows(stop_times)

    with open(f"{OUT}/calendar.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["service_id", "mon", "tue", "wed", "thu", "fri", "sat", "sun"])
        w.writerow(["평일", 1, 1, 1, 1, 1, 0, 0])
        w.writerow(["휴일", 0, 0, 0, 0, 0, 1, 1])

    print(f"stops      {len(stops):6d}")
    print(f"trips      {len(trips):6d}")
    print(f"stop_times {len(stop_times):6d}")
    print(f"warnings   {len(warnings):6d}")
    for w_ in warnings[:20]:
        print("  -", w_)


if __name__ == "__main__":
    main()
